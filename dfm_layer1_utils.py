
"""
Helper utilities for Layer 1 of the hybrid DFM + ML nowcasting project.

The code is designed to be repository-aware, vintage-aware, and careful about
the repository's first-day-of-month / first-day-of-quarter timestamp
convention. It avoids artificial month-end conversion and instead canonicalizes
time semantics with pandas PeriodIndex objects.
"""

from __future__ import annotations

import json
import math
import re
import warnings
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Mapping, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from statsmodels.tsa.statespace.dynamic_factor_mq import DynamicFactorMQ


# --------------------------------------------------------------------------------------
# Stable subset metadata from the uploaded data dictionary.
# --------------------------------------------------------------------------------------

STABLE_SUBSET_METADATA: List[Dict[str, Any]] = [
    # Real activity & income
    {"mnemonic": "RPI", "block": "real_activity_income", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "W875RX1", "block": "real_activity_income", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "INDPRO", "block": "real_activity_income", "tcode": 5, "anchor": True,  "alias": None},
    {"mnemonic": "IPFINAL", "block": "real_activity_income", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "IPCONGD", "block": "real_activity_income", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "IPBUSEQ", "block": "real_activity_income", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "IPMANSICS", "block": "real_activity_income", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "CUMFNS", "block": "real_activity_income", "tcode": 2, "anchor": False, "alias": None},

    # Labor market
    {"mnemonic": "CLF16OV", "block": "labor_market", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "UNRATE", "block": "labor_market", "tcode": 2, "anchor": False, "alias": None},
    {"mnemonic": "UEMPMEAN", "block": "labor_market", "tcode": 2, "anchor": False, "alias": None},
    {"mnemonic": "CLAIMSx", "block": "labor_market", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "PAYEMS", "block": "labor_market", "tcode": 5, "anchor": True,  "alias": None},
    {"mnemonic": "MANEMP", "block": "labor_market", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "AWHMAN", "block": "labor_market", "tcode": 1, "anchor": False, "alias": None},
    {"mnemonic": "CES3000000008", "block": "labor_market", "tcode": 6, "anchor": False, "alias": None},

    # Housing & construction
    {"mnemonic": "HOUST", "block": "housing_construction", "tcode": 4, "anchor": True,  "alias": None},
    {"mnemonic": "PERMIT", "block": "housing_construction", "tcode": 4, "anchor": False, "alias": None},
    {"mnemonic": "HOUSTNE", "block": "housing_construction", "tcode": 4, "anchor": False, "alias": None},
    {"mnemonic": "HOUSTMW", "block": "housing_construction", "tcode": 4, "anchor": False, "alias": None},
    {"mnemonic": "HOUSTS", "block": "housing_construction", "tcode": 4, "anchor": False, "alias": None},
    {"mnemonic": "HOUSTW", "block": "housing_construction", "tcode": 4, "anchor": False, "alias": None},

    # Demand / orders / inventories
    {"mnemonic": "DPCERA3M086SBEA", "block": "demand_orders_inventories", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "CMRMTSPLx", "block": "demand_orders_inventories", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "RETAILx", "block": "demand_orders_inventories", "tcode": 5, "anchor": True,  "alias": None},
    {"mnemonic": "ACOGNO", "block": "demand_orders_inventories", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "AMDMNOx", "block": "demand_orders_inventories", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "ANDENOx", "block": "demand_orders_inventories", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "AMDMUOx", "block": "demand_orders_inventories", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "BUSINVx", "block": "demand_orders_inventories", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "ISRATIOx", "block": "demand_orders_inventories", "tcode": 2, "anchor": False, "alias": None},

    # Prices & inflation
    {"mnemonic": "CPIAUCSL", "block": "prices_inflation", "tcode": 6, "anchor": True,  "alias": None},
    {"mnemonic": "CPIULFSL", "block": "prices_inflation", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "CUSR0000SAS", "block": "prices_inflation", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "CUSR0000SAC", "block": "prices_inflation", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "PCEPI", "block": "prices_inflation", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "DDURRG3M086SBEA", "block": "prices_inflation", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "DSERRG3M086SBEA", "block": "prices_inflation", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "WPSFD49207", "block": "prices_inflation", "tcode": 6, "anchor": False, "alias": "PPIFGS"},
    {"mnemonic": "OILPRICEx", "block": "prices_inflation", "tcode": 6, "anchor": False, "alias": None},

    # Financial conditions
    {"mnemonic": "FEDFUNDS", "block": "financial_conditions", "tcode": 2, "anchor": True,  "alias": None},
    {"mnemonic": "TB3MS", "block": "financial_conditions", "tcode": 2, "anchor": False, "alias": None},
    {"mnemonic": "GS10", "block": "financial_conditions", "tcode": 2, "anchor": False, "alias": None},
    {"mnemonic": "BAA", "block": "financial_conditions", "tcode": 2, "anchor": False, "alias": None},
    {"mnemonic": "T10YFFM", "block": "financial_conditions", "tcode": 1, "anchor": False, "alias": None},
    {"mnemonic": "BAAFFM", "block": "financial_conditions", "tcode": 1, "anchor": False, "alias": None},
    {"mnemonic": "BUSLOANS", "block": "financial_conditions", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "REALLN", "block": "financial_conditions", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "NONREVSL", "block": "financial_conditions", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "TWEXAFEGSMTHx", "block": "financial_conditions", "tcode": 5, "anchor": False, "alias": "TWEXMMTH"},
]

BLOCK_LABELS: Dict[str, str] = {
    "real_activity_income": "Real activity & income",
    "labor_market": "Labor market",
    "housing_construction": "Housing & construction",
    "demand_orders_inventories": "Demand / orders / inventories",
    "prices_inflation": "Prices & inflation",
    "financial_conditions": "Financial conditions",
}

BLOCK_TO_FACTOR: Dict[str, str] = {
    "real_activity_income": "real_activity",
    "labor_market": "labor",
    "housing_construction": "housing",
    "demand_orders_inventories": "demand",
    "prices_inflation": "prices",
    "financial_conditions": "financial",
}

MNEMONIC_CROSSWALK: Dict[str, str] = {
    "PPIFGS": "WPSFD49207",
    "TWEXMMTH": "TWEXAFEGSMTHx",
}

TARGET_FILE_PRIORITY = [
    "routputMvQd.xlsx",
    "ROUTPUTQvQd.xlsx",
    "routput_first_second_third.xlsx",
    "GDPplus_Vintages.xlsx",
    "meanGrowth.xlsx",
    "medianGrowth.xlsx",
]


@dataclass
class ProtocolConfig:
    repo_root: str
    output_dir: str
    benchmark_start_quarter: str = "2000Q1"
    panel_mode: str = "stable"  # "stable" is the reproducible block-structured default.
    truth_main: str = "third_release"
    truth_robustness: Tuple[str, ...] = ("latest_rtdsm", "gdpplus")
    candidate_factor_orders: Tuple[int, ...] = (1, 2)
    fixed_factor_order: Optional[int] = 1
    select_factor_order_per_vintage: bool = False
    idiosyncratic_ar1: bool = True
    em_maxiter: int = 100
    em_tolerance: float = 1e-6
    warm_start_from_previous_vintage: bool = True
    export_same_tau_residual_lags: bool = False
    state_standardize: bool = True
    state_center: bool = True
    state_scale_floor: float = 1e-6
    vintage_limit: Optional[int] = None
    min_monthly_obs: int = 24
    force_refit: bool = True


# --------------------------------------------------------------------------------------
# Generic helpers
# --------------------------------------------------------------------------------------

def stable_subset_metadata() -> pd.DataFrame:
    out = pd.DataFrame(STABLE_SUBSET_METADATA)
    out["block_label"] = out["block"].map(BLOCK_LABELS)
    out["factor_name"] = out["block"].map(BLOCK_TO_FACTOR)
    return out


def normalize_column_name(value: Any) -> str:
    value = "" if value is None else str(value)
    value = value.strip()
    value = re.sub(r"\s+", "_", value)
    return value


def locate_repo_root(start: Optional[Path] = None) -> Path:
    """
    Search upwards for a repository root containing data/raw.
    """
    if start is None:
        start = Path.cwd()
    start = start.resolve()
    candidates = [start] + list(start.parents)
    for root in candidates:
        if (root / "data" / "raw").exists():
            return root
    raise FileNotFoundError(
        "Could not locate repository root containing data/raw. "
        "Place the notebook in the repository or set REPO_ROOT manually."
    )


def ensure_directory(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def to_month_period_from_filename(year: int, month: int) -> pd.Period:
    return pd.Period(f"{int(year):04d}-{int(month):02d}", freq="M")


def parse_vintage_from_filename(filename: str) -> Optional[pd.Period]:
    """
    Parse repository vintage markers without forcing month-end timestamps.

    The repository mixes ISO-like filenames (``YYYY-MM-DD.csv``), plain monthly
    stamps (``YYYY-MM.csv``), and FRED historical vintages such as
    ``FRED-MD_2024m03.csv`` or ``FRED-QD_2019m1.csv``. The parser must accept
    both one-digit and two-digit month codes so that internal vintage schedules
    are complete and pseudo-real-time gaps are not created mechanically.
    """
    name = Path(filename).name

    patterns = [
        r"(?P<y>\d{4})-(?P<m>\d{2})-(?P<d>\d{2})\.csv$",
        r"(?P<y>\d{4})-(?P<m>\d{2})(?:-(?:MD|QD))?\.csv$",
        r"FRED-(?:MD|QD)_(?P<y>\d{4})m(?P<m>\d{1,2})\.csv$",
    ]
    for pattern in patterns:
        match = re.search(pattern, name, flags=re.IGNORECASE)
        if match:
            y = int(match.group("y"))
            m = int(match.group("m"))
            return to_month_period_from_filename(y, m)

    return None


def classify_repo_file(path: Path) -> str:
    parts = tuple(path.parts)
    name = path.name
    if "sample" in parts:
        return "sample_csv"
    if "FRED-MD-MONTHLY" in parts:
        return "fred_md_current"
    if "Historical-vintages-of-FRED-MD-2015-01-to-2024-12" in parts:
        return "fred_md_hist_2015_2024"
    if "Historical FRED-MD Vintages Final" in parts:
        return "fred_md_hist_legacy"
    if "FRED-QD-QUARTERLY" in parts:
        return "fred_qd_current"
    if "Historical vintages of FRED-QD 2018-05 to 2024-12" in parts:
        return "fred_qd_hist_2018_2024"
    if name.lower().endswith(".xlsx"):
        return "excel_workbook"
    return "other"


def build_repo_catalog(repo_root: Path) -> pd.DataFrame:
    """
    Recursively audit repository files. Uses filename-based vintage parsing and
    keeps canonical vintage markers as monthly Periods.
    """
    rows: List[Dict[str, Any]] = []
    for path in sorted(repo_root.rglob("*")):
        if path.is_dir():
            continue
        rel = path.relative_to(repo_root)
        vintage = parse_vintage_from_filename(path.name)
        rows.append(
            {
                "path": str(rel),
                "stem": path.stem,
                "suffix": path.suffix.lower(),
                "group": classify_repo_file(rel),
                "size_bytes": path.stat().st_size,
                "vintage_period": vintage,
                "vintage_timestamp_start": vintage.to_timestamp() if vintage is not None else pd.NaT,
                "name": path.name,
            }
        )
    catalog = pd.DataFrame(rows)
    if not catalog.empty:
        catalog["size_mb"] = catalog["size_bytes"] / 1024**2
    return catalog


def choose_canonical_md_manifest(catalog: pd.DataFrame) -> pd.DataFrame:
    if catalog.empty:
        return catalog.copy()

    md = catalog[catalog["group"].isin({"fred_md_current", "fred_md_hist_2015_2024", "fred_md_hist_legacy"})].copy()
    md = md[md["vintage_period"].notna()].copy()
    if md.empty:
        return md

    priority = {
        "fred_md_hist_legacy": 1,
        "fred_md_hist_2015_2024": 2,
        "fred_md_current": 3,
    }
    md["priority"] = md["group"].map(priority)
    md = md.sort_values(["vintage_period", "priority", "path"]).drop_duplicates("vintage_period", keep="last")
    return md.sort_values("vintage_period").reset_index(drop=True)


def choose_canonical_qd_manifest(catalog: pd.DataFrame) -> pd.DataFrame:
    if catalog.empty:
        return catalog.copy()

    qd = catalog[catalog["group"].isin({"fred_qd_current", "fred_qd_hist_2018_2024"})].copy()
    qd = qd[qd["vintage_period"].notna()].copy()
    if qd.empty:
        return qd

    priority = {
        "fred_qd_hist_2018_2024": 1,
        "fred_qd_current": 2,
    }
    qd["priority"] = qd["group"].map(priority)
    qd = qd.sort_values(["vintage_period", "priority", "path"]).drop_duplicates("vintage_period", keep="last")
    return qd.sort_values("vintage_period").reset_index(drop=True)


def summarize_manifest(manifest: pd.DataFrame, label: str) -> pd.DataFrame:
    if manifest.empty:
        return pd.DataFrame([{"label": label, "count": 0, "min_vintage": pd.NaT, "max_vintage": pd.NaT}])
    return pd.DataFrame(
        [
            {
                "label": label,
                "count": len(manifest),
                "min_vintage": manifest["vintage_period"].min(),
                "max_vintage": manifest["vintage_period"].max(),
                "min_path": manifest.loc[manifest["vintage_period"].idxmin(), "path"],
                "max_path": manifest.loc[manifest["vintage_period"].idxmax(), "path"],
            }
        ]
    )


# --------------------------------------------------------------------------------------
# Timestamp / period parsing that respects first-of-month / first-of-quarter storage.
# --------------------------------------------------------------------------------------


_QUARTER_PATTERNS = [
    re.compile(r"^\s*(?P<y>\d{4})\s*[:\-]?\s*[Qq](?P<q>[1-4])\s*$"),
    re.compile(r"^\s*(?P<y>\d{4})\s*[Qq](?P<q>[1-4])\s*$"),
]

_QUARTER_NUMBER_PATTERN = re.compile(r"^\s*(?P<y>\d{4})\s*[:\-]\s*(?P<q>0?[1-4])\s*$")
_YEARMONTH_CODE_PATTERN = re.compile(r"^\s*(?P<y>\d{4})[mM](?P<m>\d{2})\s*$")
_YEARMONTH_SEP_PATTERN = re.compile(r"^\s*(?P<y>\d{4})\s*[:\-/]\s*(?P<m>0?[1-9]|1[0-2])\s*$")
_RTDMS_MONTHLY_VINTAGE_PATTERN = re.compile(r"^\s*ROUTPUT(?P<yy>\d{2})[mM](?P<m>\d{1,2})\s*$", flags=re.IGNORECASE)
_RTDMS_QUARTERLY_VINTAGE_PATTERN = re.compile(r"^\s*ROUTPUT(?P<yy>\d{2})[qQ](?P<q>[1-4])\s*$", flags=re.IGNORECASE)
_GDPPLUS_RELEASE_PATTERN = re.compile(r"^\s*GDPPLUS[_\-]?(?P<mm>\d{2})(?P<dd>\d{2})(?P<yy>\d{2})\s*$", flags=re.IGNORECASE)


def _two_digit_year_to_four_digit(yy: int) -> int:
    yy = int(yy)
    return 1900 + yy if yy >= 50 else 2000 + yy


def _quarter_number_to_period(year: int, quarter: int) -> pd.Period:
    return pd.Period(f"{int(year):04d}Q{int(quarter)}", freq="Q")


def _quarter_vintage_to_month_period(year: int, quarter: int) -> pd.Period:
    month = {1: 2, 2: 5, 3: 8, 4: 11}[int(quarter)]
    return pd.Period(f"{int(year):04d}-{int(month):02d}", freq="M")


def _parse_routput_vintage_period(text: str) -> Optional[pd.Period]:
    match = _RTDMS_MONTHLY_VINTAGE_PATTERN.match(text)
    if match:
        year = _two_digit_year_to_four_digit(int(match.group("yy")))
        month = int(match.group("m"))
        return pd.Period(f"{year:04d}-{month:02d}", freq="M")

    match = _RTDMS_QUARTERLY_VINTAGE_PATTERN.match(text)
    if match:
        year = _two_digit_year_to_four_digit(int(match.group("yy")))
        quarter = int(match.group("q"))
        return _quarter_vintage_to_month_period(year, quarter)

    return None


def _parse_gdpplus_release_timestamp(text: str) -> Optional[pd.Timestamp]:
    match = _GDPPLUS_RELEASE_PATTERN.match(text)
    if not match:
        return None
    month = int(match.group("mm"))
    day = int(match.group("dd"))
    year = _two_digit_year_to_four_digit(int(match.group("yy")))
    try:
        return pd.Timestamp(year=year, month=month, day=day)
    except Exception:
        return None


def _looks_like_explicit_quarter_label(text: str) -> bool:
    return any(pattern.match(text) for pattern in _QUARTER_PATTERNS) or _QUARTER_NUMBER_PATTERN.match(text) is not None


def _looks_like_explicit_month_label(text: str) -> bool:
    return (
        _YEARMONTH_CODE_PATTERN.match(text) is not None
        or _RTDMS_MONTHLY_VINTAGE_PATTERN.match(text) is not None
        or _RTDMS_QUARTERLY_VINTAGE_PATTERN.match(text) is not None
        or _GDPPLUS_RELEASE_PATTERN.match(text) is not None
    )


def parse_periodish(value: Any, freq_hint: Optional[str] = None) -> Optional[pd.Period]:
    """
    Parse month-like or quarter-like values to a pandas Period without shifting
    to month-end. The function intentionally preserves first-of-period semantics.

    Important design rule for this repository:
    - plain first-day timestamps like 2026-01-01 default to monthly semantics
      unless the caller explicitly asks for quarterly parsing.
    """
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, pd.Period):
        if freq_hint == "Q":
            return value.asfreq("Q")
        if freq_hint == "M":
            return value.asfreq("M")
        return value
    if isinstance(value, pd.Timestamp):
        if freq_hint == "Q":
            return value.to_period("Q")
        return value.to_period("M")

    text = str(value).strip()
    if text == "":
        return None

    routput_period = _parse_routput_vintage_period(text)
    if routput_period is not None:
        if freq_hint == "Q":
            return routput_period.asfreq("Q")
        return routput_period

    gdpplus_release_ts = _parse_gdpplus_release_timestamp(text)
    if gdpplus_release_ts is not None:
        if freq_hint == "Q":
            return gdpplus_release_ts.to_period("Q")
        return gdpplus_release_ts.to_period("M")

    for pattern in _QUARTER_PATTERNS:
        match = pattern.match(text)
        if match:
            return _quarter_number_to_period(int(match.group("y")), int(match.group("q")))

    match = _QUARTER_NUMBER_PATTERN.match(text)
    if match:
        return _quarter_number_to_period(int(match.group("y")), int(match.group("q")))

    match = _YEARMONTH_CODE_PATTERN.match(text)
    if match:
        year = int(match.group("y"))
        month = int(match.group("m"))
        if freq_hint == "Q" and month in (1, 4, 7, 10):
            return _quarter_number_to_period(year, ((month - 1) // 3) + 1)
        return pd.Period(f"{year:04d}-{month:02d}", freq="M")

    match = _YEARMONTH_SEP_PATTERN.match(text)
    if match:
        year = int(match.group("y"))
        month = int(match.group("m"))
        if freq_hint == "Q" and month in (1, 4, 7, 10):
            return _quarter_number_to_period(year, ((month - 1) // 3) + 1)
        return pd.Period(f"{year:04d}-{month:02d}", freq="M")

    try:
        ts = pd.to_datetime(text, errors="raise")
        if freq_hint == "Q":
            return ts.to_period("Q")
        if freq_hint == "M":
            return ts.to_period("M")
        return ts.to_period("M")
    except Exception:
        return None

def parse_timestamp_series(values: pd.Series, freq_hint: Optional[str] = None) -> pd.Series:
    """
    Parse to Timestamp series while respecting repository first-day conventions.
    """
    periods = values.map(lambda x: parse_periodish(x, freq_hint=freq_hint))
    out = pd.Series(pd.NaT, index=values.index, dtype="datetime64[ns]")
    for idx, p in periods.items():
        if p is not None and not pd.isna(p):
            out.loc[idx] = p.to_timestamp()
    return out



def infer_period_frequency_from_values(values: Sequence[Any]) -> Optional[str]:
    """
    Infer whether a collection of labels is monthly or quarterly while respecting
    repository first-day semantics and the RTDSM / GDPplus workbook conventions.
    """
    explicit_quarter_hits = 0
    explicit_month_hits = 0
    timestamps: List[pd.Timestamp] = []

    for value in values:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            continue

        text = str(value).strip()
        if text == "":
            continue

        if _looks_like_explicit_quarter_label(text):
            explicit_quarter_hits += 1
            p = parse_periodish(text, freq_hint="Q")
            if p is not None:
                timestamps.append(p.to_timestamp())
            continue

        if _looks_like_explicit_month_label(text):
            explicit_month_hits += 1
            p = parse_periodish(text, freq_hint="M")
            if p is not None:
                timestamps.append(p.to_timestamp())
            continue

        try:
            ts = pd.to_datetime(value, errors="raise")
            timestamps.append(pd.Timestamp(ts))
        except Exception:
            continue

    if explicit_quarter_hits and explicit_quarter_hits >= max(3, explicit_month_hits + 1):
        return "Q"
    if explicit_month_hits and explicit_month_hits >= max(3, explicit_quarter_hits + 1):
        return "M"

    if not timestamps:
        return None

    ordinals = np.array([ts.year * 12 + ts.month for ts in timestamps], dtype=int)
    ordinals = np.unique(np.sort(ordinals))
    quarter_like_month_sets = ({1, 4, 7, 10}, {2, 5, 8, 11}, {3, 6, 9, 12})

    if len(ordinals) >= 2:
        diffs = np.diff(ordinals)
        if len(diffs) > 0:
            month_set = set((o - 1) % 12 + 1 for o in ordinals)
            if np.mean(diffs == 3) > 0.8 and any(month_set.issubset(s) for s in quarter_like_month_sets):
                return "Q"
            if np.mean(diffs == 1) > 0.8:
                return "M"

    months = {(o - 1) % 12 + 1 for o in ordinals}
    if any(months.issubset(s) for s in quarter_like_month_sets) and len(ordinals) >= 4:
        return "Q"
    return "M"

# --------------------------------------------------------------------------------------
# CSV inspection and FRED snapshot loading
# --------------------------------------------------------------------------------------

def detect_date_column(df: pd.DataFrame) -> str:
    preferred = ["sasdate", "date", "observation_date", "DATE", "SASDATE"]
    for col in df.columns:
        if str(col) in preferred:
            return col
    # otherwise use the column with the highest parse success among the first few columns
    best_col = df.columns[0]
    best_score = -1.0
    for col in list(df.columns[: min(5, len(df.columns))]):
        score = pd.Series(df[col]).map(lambda x: parse_periodish(x) is not None).mean()
        if score > best_score:
            best_score = float(score)
            best_col = col
    return best_col


def _coalesce_duplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    if not df.columns.duplicated().any():
        return df
    pieces = []
    for col in pd.Index(df.columns).unique():
        sub = df.loc[:, df.columns == col]
        if sub.shape[1] == 1:
            pieces.append(sub.iloc[:, 0].rename(col))
        else:
            combined = sub.bfill(axis=1).iloc[:, 0].rename(col)
            pieces.append(combined)
    out = pd.concat(pieces, axis=1)
    out = out.loc[:, ~out.columns.duplicated()]
    return out


def _apply_mnemonic_crosswalk_to_columns(columns: Sequence[Any]) -> List[str]:
    mapped = []
    for col in columns:
        text = str(col)
        mapped.append(MNEMONIC_CROSSWALK.get(text, text))
    return mapped


def _leading_metadata_row_info(df: pd.DataFrame, date_col: str) -> Dict[str, Any]:
    metadata_rows: List[int] = []
    tcode_row_idx: Optional[int] = None
    for i in range(min(5, len(df))):
        label = str(df.iloc[i][date_col]).strip().lower()
        numeric_share = (
            pd.to_numeric(df.iloc[i].drop(labels=[date_col]), errors="coerce")
            .between(1, 7, inclusive="both")
            .mean()
        )
        is_metadata = (
            label.startswith("transform")
            or label.startswith("tcode")
            or label.startswith("factor")
            or numeric_share > 0.6
        )
        if is_metadata:
            metadata_rows.append(i)
            if (label.startswith("transform") or label.startswith("tcode") or numeric_share > 0.6) and tcode_row_idx is None:
                tcode_row_idx = i
        else:
            break
    return {
        "metadata_rows": metadata_rows,
        "metadata_row_count": len(metadata_rows),
        "tcode_row_idx": tcode_row_idx,
        "has_tcode_row": tcode_row_idx is not None,
    }


def inspect_csv_schema(path: Path, n_preview_rows: int = 6) -> Dict[str, Any]:
    """
    Read a CSV using the most likely FRED-style schema and return a human-readable inspection.
    """
    df = pd.read_csv(path, low_memory=False)
    date_col = detect_date_column(df)
    meta = _leading_metadata_row_info(df, date_col)

    preview_df = df.head(n_preview_rows).copy()
    raw_date_series = df[date_col].iloc[meta["metadata_row_count"]:].copy()
    inferred_freq = infer_period_frequency_from_values(raw_date_series.head(50).tolist())
    parsed_ts = parse_timestamp_series(raw_date_series, freq_hint=inferred_freq)
    parsed_ts = parsed_ts.dropna()

    out = {
        "path": str(path),
        "n_rows": len(df),
        "n_cols": df.shape[1],
        "date_col": str(date_col),
        "has_tcode_row": bool(meta["has_tcode_row"]),
        "metadata_row_count": int(meta["metadata_row_count"]),
        "preview": preview_df,
        "min_timestamp": parsed_ts.min() if not parsed_ts.empty else pd.NaT,
        "max_timestamp": parsed_ts.max() if not parsed_ts.empty else pd.NaT,
        "inferred_freq": inferred_freq,
        "columns_preview": list(map(str, df.columns[: min(15, len(df.columns))])),
    }
    return out


def load_fred_snapshot(path: Path, freq_hint: Optional[str] = None) -> Dict[str, Any]:
    """
    Load a FRED-MD / FRED-QD style vintage snapshot.

    The function preserves the repository's first-day timestamp convention by
    creating a PeriodIndex (M or Q) from the raw date column instead of
    artificially shifting dates to month-end.
    """
    df = pd.read_csv(path, low_memory=False)
    df.columns = [normalize_column_name(c) for c in df.columns]
    date_col = detect_date_column(df)
    meta = _leading_metadata_row_info(df, date_col)

    tcodes = pd.Series(dtype="Int64")
    if meta["tcode_row_idx"] is not None:
        raw_tcodes = pd.to_numeric(df.iloc[meta["tcode_row_idx"]].drop(labels=[date_col]), errors="coerce")
        tcodes = raw_tcodes.astype("Int64")

    if meta["metadata_row_count"] > 0:
        df = df.iloc[meta["metadata_row_count"]:].copy()

    df.columns = [normalize_column_name(c) for c in _apply_mnemonic_crosswalk_to_columns(df.columns)]
    date_col = normalize_column_name(MNEMONIC_CROSSWALK.get(date_col, date_col))

    # If freq_hint is missing, infer from the raw date column after stripping metadata rows.
    inferred_freq = freq_hint or infer_period_frequency_from_values(df[date_col].head(100).tolist()) or "M"
    date_ts = parse_timestamp_series(df[date_col], freq_hint=inferred_freq)

    numeric = df.drop(columns=[date_col]).apply(pd.to_numeric, errors="coerce")
    numeric.columns = _apply_mnemonic_crosswalk_to_columns(numeric.columns)
    numeric = _coalesce_duplicate_columns(numeric)

    if inferred_freq.upper().startswith("Q"):
        index = date_ts.dt.to_period("Q")
    else:
        index = date_ts.dt.to_period("M")

    valid = ~pd.isna(index)
    numeric = numeric.loc[valid].copy()
    index = index[valid]
    date_ts = date_ts.loc[valid]

    numeric.index = index
    numeric.index.name = "period"

    tcodes.index = _apply_mnemonic_crosswalk_to_columns(tcodes.index)
    tcodes = tcodes[~pd.Index(tcodes.index).duplicated(keep="last")]

    return {
        "path": str(path),
        "date_col": date_col,
        "tcode_row_present": bool(meta["has_tcode_row"]),
        "metadata_row_count": int(meta["metadata_row_count"]),
        "tcodes": tcodes,
        "data": numeric.sort_index(),
        "raw_timestamps": pd.Series(date_ts.values, index=index, name="timestamp_start").sort_index(),
        "freq": inferred_freq.upper()[0],
    }


# --------------------------------------------------------------------------------------
# Transformation logic
# --------------------------------------------------------------------------------------

def apply_tcode(series: pd.Series, tcode: int) -> pd.Series:
    """
    Apply official FRED-MD / FRED-QD transformation code.
    """
    s = pd.to_numeric(series, errors="coerce").astype(float)

    if tcode == 1:
        out = s
    elif tcode == 2:
        out = s.diff()
    elif tcode == 3:
        out = s.diff().diff()
    elif tcode == 4:
        out = np.log(s.where(s > 0))
    elif tcode == 5:
        out = np.log(s.where(s > 0)).diff()
    elif tcode == 6:
        out = np.log(s.where(s > 0)).diff().diff()
    elif tcode == 7:
        out = (s / s.shift(1) - 1.0).diff()
    else:
        raise ValueError(f"Unsupported tcode {tcode!r}")
    out.name = series.name
    return out


def apply_tcodes_to_snapshot(snapshot: Dict[str, Any], tcode_override: Optional[Mapping[str, int]] = None) -> pd.DataFrame:
    data = snapshot["data"].copy()
    tcodes = snapshot["tcodes"].copy()
    if tcode_override is not None:
        for key, value in tcode_override.items():
            tcodes.loc[key] = int(value)

    transformed = {}
    for col in data.columns:
        tcode = tcodes.get(col, pd.NA)
        if pd.isna(tcode):
            transformed[col] = data[col]
        else:
            transformed[col] = apply_tcode(data[col], int(tcode))
    out = pd.DataFrame(transformed, index=data.index)
    return out


# --------------------------------------------------------------------------------------
# Excel inspection and flexible extraction
# --------------------------------------------------------------------------------------

def inspect_excel_workbook(path: Path, preview_rows: int = 6, preview_cols: int = 8) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: List[Dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        preview: List[List[Any]] = []
        for row in ws.iter_rows(min_row=1, max_row=preview_rows, max_col=preview_cols, values_only=True):
            preview.append(list(row))
        rows.append(
            {
                "sheet_name": sheet_name,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "preview_top_left": preview,
            }
        )
    return pd.DataFrame(rows)


def _read_sheet_as_values(path: Path, sheet_name: str, max_rows: Optional[int] = None, max_cols: Optional[int] = None) -> List[List[Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows: List[List[Any]] = []
    kwargs: Dict[str, Any] = {"values_only": True}
    if max_rows is not None:
        kwargs["max_row"] = max_rows
    if max_cols is not None:
        kwargs["max_col"] = max_cols
    for row in ws.iter_rows(**kwargs):
        rows.append(list(row))
    return rows


def _trim_2d(values: List[List[Any]]) -> List[List[Any]]:
    if not values:
        return values
    n_cols = max(len(r) for r in values)
    arr = [list(r) + [None] * (n_cols - len(r)) for r in values]

    # Trim empty bottom rows
    while arr and all(v in (None, "") or (isinstance(v, float) and math.isnan(v)) for v in arr[-1]):
        arr.pop()

    if not arr:
        return arr

    # Trim empty right columns
    last_nonempty = -1
    for j in range(n_cols):
        if any(row[j] not in (None, "") and not (isinstance(row[j], float) and math.isnan(row[j])) for row in arr):
            last_nonempty = j
    if last_nonempty >= 0:
        arr = [row[: last_nonempty + 1] for row in arr]
    return arr


def _score_matrix_candidate(values: List[List[Any]], header_row: int, index_col: int) -> Dict[str, Any]:
    trimmed = _trim_2d(values)
    if not trimmed or header_row >= len(trimmed):
        return {"score": -1}

    n_cols = max(len(r) for r in trimmed)
    if index_col >= n_cols:
        return {"score": -1}

    row_labels = trimmed[header_row][index_col + 1 :]
    col_labels = [trimmed[i][index_col] if index_col < len(trimmed[i]) else None for i in range(header_row + 1, len(trimmed))]

    parsed_rows = [parse_periodish(v) for v in row_labels]
    parsed_cols = [parse_periodish(v) for v in col_labels]

    row_count = sum(v is not None for v in parsed_rows)
    col_count = sum(v is not None for v in parsed_cols)

    if row_count < 3 or col_count < 3:
        return {"score": -1}

    body = []
    for i in range(header_row + 1, len(trimmed)):
        row = trimmed[i]
        numeric_row = row[index_col + 1 :]
        body.extend([pd.to_numeric(x, errors="coerce") for x in numeric_row if x is not None])
    body = pd.Series(body)
    numeric_share = body.notna().mean() if len(body) else 0.0

    row_freq = infer_period_frequency_from_values(row_labels)
    col_freq = infer_period_frequency_from_values(col_labels)

    score = row_count + col_count + 10.0 * numeric_share
    return {
        "score": score,
        "header_row": header_row,
        "index_col": index_col,
        "row_freq": row_freq,
        "col_freq": col_freq,
        "row_labels_parsed": parsed_rows,
        "col_labels_parsed": parsed_cols,
    }



def extract_best_period_matrix(path: Path, required_col_freq: Optional[str] = None) -> Dict[str, Any]:
    """
    Search workbook sheets for a vintage-by-observation matrix where either rows
    or columns look like monthly/quarterly periods.

    This is used for RTDSM-style ROUTPUT and GDPplus workbooks.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    candidates: List[Dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        values = _read_sheet_as_values(path, sheet_name)
        values = _trim_2d(values)
        if not values:
            continue

        n_rows = len(values)
        n_cols = max(len(r) for r in values)
        for header_row in range(min(8, n_rows)):
            for index_col in range(min(6, n_cols)):
                cand = _score_matrix_candidate(values, header_row, index_col)
                if cand["score"] < 0:
                    continue
                cand["sheet_name"] = sheet_name
                cand["values"] = values
                if required_col_freq is not None and cand.get("row_freq") != required_col_freq and cand.get("col_freq") != required_col_freq:
                    continue
                candidates.append(cand)

    if not candidates:
        raise ValueError(f"Could not find a date-like matrix in workbook: {path}")

    def ranking_key(c: Dict[str, Any]) -> Tuple[int, float]:
        has_required = int(required_col_freq is None or c.get("row_freq") == required_col_freq or c.get("col_freq") == required_col_freq)
        return (has_required, float(c["score"]))

    best = sorted(candidates, key=ranking_key, reverse=True)[0]
    values = best["values"]
    header_row = int(best["header_row"])
    index_col = int(best["index_col"])

    raw_col_labels = values[header_row][index_col + 1 :]
    raw_row_labels = [values[i][index_col] if index_col < len(values[i]) else None for i in range(header_row + 1, len(values))]
    row_freq = infer_period_frequency_from_values(raw_row_labels)
    col_freq = infer_period_frequency_from_values(raw_col_labels)
    row_periods = [parse_periodish(v, freq_hint=row_freq) for v in raw_row_labels]
    col_periods = [parse_periodish(v, freq_hint=col_freq) for v in raw_col_labels]

    body_rows: List[List[float]] = []
    for i in range(header_row + 1, len(values)):
        row = values[i] + [None] * (len(raw_col_labels) + index_col + 1 - len(values[i]))
        body_rows.append([pd.to_numeric(x, errors="coerce") for x in row[index_col + 1 : index_col + 1 + len(raw_col_labels)]])

    body_df = pd.DataFrame(body_rows)
    row_keep = ~body_df.isna().all(axis=1)
    col_keep = ~body_df.isna().all(axis=0)

    kept_row_periods = [p for p, keep in zip(row_periods, row_keep) if bool(keep)]
    kept_col_periods = [p for p, keep in zip(col_periods, col_keep) if bool(keep)]
    kept_raw_row_labels = [v for v, keep in zip(raw_row_labels, row_keep) if bool(keep)]
    kept_raw_col_labels = [v for v, keep in zip(raw_col_labels, col_keep) if bool(keep)]

    matrix = pd.DataFrame(
        body_df.loc[row_keep, col_keep].to_numpy(),
        index=kept_row_periods,
        columns=kept_col_periods,
    )

    return {
        "sheet_name": best["sheet_name"],
        "header_row": header_row,
        "index_col": index_col,
        "row_freq": row_freq,
        "col_freq": col_freq,
        "matrix": matrix,
        "raw_row_labels": raw_row_labels,
        "raw_col_labels": raw_col_labels,
        "kept_raw_row_labels": kept_raw_row_labels,
        "kept_raw_col_labels": kept_raw_col_labels,
    }

def melt_vintage_matrix(matrix_info: Dict[str, Any], value_name: str = "value") -> pd.DataFrame:
    matrix = matrix_info["matrix"].copy()
    row_freq = matrix_info["row_freq"]
    col_freq = matrix_info["col_freq"]

    if row_freq not in {"M", "Q"} and col_freq not in {"M", "Q"}:
        raise ValueError("Matrix does not expose monthly/quarterly period axes.")

    if row_freq in {"M", "Q"} and col_freq == "Q":
        long = (
            matrix.stack(dropna=False)
            .rename(value_name)
            .rename_axis(index=["vintage_period", "obs_period"])
            .reset_index()
        )
        long["vintage_freq"] = row_freq
        long["obs_freq"] = col_freq
        return long

    if col_freq in {"M", "Q"} and row_freq == "Q":
        long = (
            matrix.T.stack(dropna=False)
            .rename(value_name)
            .rename_axis(index=["vintage_period", "obs_period"])
            .reset_index()
        )
        long["vintage_freq"] = col_freq
        long["obs_freq"] = row_freq
        return long

    raise ValueError("Could not orient matrix as vintage x observed-period.")


def load_routput_vintage_history(path: Path) -> pd.DataFrame:
    matrix_info = extract_best_period_matrix(path, required_col_freq="Q")
    long = melt_vintage_matrix(matrix_info, value_name="level")
    long = long[long["obs_freq"] == "Q"].copy()
    long = long.dropna(subset=["vintage_period", "obs_period"])
    long["vintage_timestamp_start"] = long["vintage_period"].map(lambda p: p.to_timestamp() if isinstance(p, pd.Period) else pd.NaT)
    long["quarter_timestamp_start"] = long["obs_period"].map(lambda p: p.to_timestamp() if isinstance(p, pd.Period) else pd.NaT)

    # Growth computed from levels within vintage.
    long = long.sort_values(["vintage_period", "obs_period"]).reset_index(drop=True)
    long["gdp_growth_annualized"] = (
        long.groupby("vintage_period")["level"]
        .transform(lambda s: 400.0 * np.log(s / s.shift(1)))
    )
    return long



def _candidate_excel_tables(path: Path) -> Iterator[Tuple[str, int, pd.DataFrame]]:
    xl = pd.ExcelFile(path)
    for sheet in xl.sheet_names:
        for header in range(0, 10):
            try:
                df = pd.read_excel(path, sheet_name=sheet, header=header)
            except Exception:
                continue
            if df is None or df.empty:
                continue
            df = df.copy()
            df.columns = [normalize_column_name(c).lower() for c in df.columns]
            yield sheet, header, df

def _find_period_column(df: pd.DataFrame) -> Optional[str]:
    best_col = None
    best_score = 0.0
    for col in df.columns:
        score = df[col].map(lambda x: parse_periodish(x, freq_hint="Q") is not None).mean()
        if score > best_score:
            best_score = float(score)
            best_col = col
    return best_col if best_score > 0.5 else None



def load_release_truth_table(path: Path) -> pd.DataFrame:
    """
    Load the RTDSM first/second/third workbook into a tidy table.

    Expected output columns:
        quarter, first_release, second_release, third_release, latest
    """
    alias_map = {
        "quarter": ["date", "quarter"],
        "first_release": ["first", "first_release"],
        "second_release": ["second", "second_release"],
        "third_release": ["third", "third_release"],
        "latest": ["most_recent", "mostrecent", "latest"],
    }

    best_df: Optional[pd.DataFrame] = None
    best_period_col: Optional[str] = None
    best_score = -np.inf

    for _, header, df in _candidate_excel_tables(path):
        cols = list(df.columns)
        period_candidates = [c for c in cols if c in alias_map["quarter"]]
        period_col = period_candidates[0] if period_candidates else _find_period_column(df)
        if period_col is None:
            continue

        found: Dict[str, str] = {}
        score = 0.0

        for canonical, aliases in alias_map.items():
            if canonical == "quarter":
                continue
            for alias in aliases:
                exact = [c for c in cols if c == alias]
                prefix = [c for c in cols if c.startswith(alias)]
                matches = exact or prefix
                if matches:
                    found[canonical] = matches[0]
                    break

        if "third_release" in found:
            score += 10.0
        if "second_release" in found:
            score += 6.0
        if "first_release" in found:
            score += 6.0
        if "latest" in found:
            score += 4.0

        score += 2.0 * float(period_col in {"date", "quarter"})
        score -= 0.25 * float(header)

        if score > best_score:
            best_score = score
            best_df = df
            best_period_col = period_col

    if best_df is None or best_period_col is None:
        raise ValueError(f"Could not detect a release truth table in workbook: {path}")

    out = pd.DataFrame()
    out["quarter"] = best_df[best_period_col].map(lambda x: parse_periodish(x, freq_hint="Q"))
    for canonical, aliases in alias_map.items():
        if canonical == "quarter":
            continue
        matches = [c for c in best_df.columns if any(c == alias or c.startswith(alias) for alias in aliases)]
        if matches:
            out[canonical] = pd.to_numeric(best_df[matches[0]], errors="coerce")

    value_cols = [c for c in out.columns if c != "quarter"]
    out = out.dropna(subset=["quarter"])
    if value_cols:
        out = out.dropna(subset=value_cols, how="all")
    return out.reset_index(drop=True)


def load_gdpplus_latest_table(path: Path, value_label: str = "gdpplus") -> pd.DataFrame:
    """
    Parse GDPplus_Vintages.xlsx, which is a wide quarter-by-vintage matrix with
    release headers such as GDPPLUS_082913.
    """
    matrix_info = extract_best_period_matrix(path, required_col_freq="Q")
    matrix = matrix_info["matrix"].copy()
    row_freq = matrix_info["row_freq"]
    col_freq = matrix_info["col_freq"]

    if row_freq == "Q":
        quarter_index = pd.PeriodIndex(matrix.index, freq="Q")
        raw_release_labels = matrix_info.get("kept_raw_col_labels", list(matrix.columns))
        value_df = pd.DataFrame(matrix.to_numpy(), index=quarter_index, columns=range(matrix.shape[1]))
    elif col_freq == "Q":
        quarter_index = pd.PeriodIndex(matrix.columns, freq="Q")
        raw_release_labels = matrix_info.get("kept_raw_row_labels", list(matrix.index))
        value_df = pd.DataFrame(matrix.T.to_numpy(), index=quarter_index, columns=range(matrix.shape[0]))
    else:
        raise ValueError(f"Could not orient GDPplus matrix in workbook: {path}")

    release_dates: List[pd.Timestamp] = []
    for raw in raw_release_labels:
        ts = _parse_gdpplus_release_timestamp(str(raw))
        if ts is None:
            p = parse_periodish(raw, freq_hint="M")
            ts = p.to_timestamp() if p is not None else pd.NaT
        release_dates.append(ts)

    value_df.columns = release_dates
    long = (
        value_df.stack(dropna=False)
        .rename(value_label)
        .rename_axis(index=["quarter", "release_date"])
        .reset_index()
    )
    long = long.dropna(subset=["quarter"])
    long["release_date"] = pd.to_datetime(long["release_date"], errors="coerce")
    long = long.dropna(subset=["release_date"])
    long = long.dropna(subset=[value_label])
    long = long.sort_values(["quarter", "release_date"]).reset_index(drop=True)
    latest = long.groupby("quarter", as_index=False).tail(1).reset_index(drop=True)
    latest["latest_release_period"] = latest["release_date"].dt.to_period("M")
    return latest[["quarter", value_label, "release_date", "latest_release_period"]]


def load_spf_growth_benchmark(path: Path, value_label: str) -> pd.DataFrame:
    """
    Parse MeanGrowth.xlsx / MedianGrowth.xlsx using the RGDP worksheet.

    The relevant sheet stores survey-quarter identifiers in YEAR and QUARTER and
    the current-quarter real GDP growth benchmark in the left-most drgdp* column
    (for the current files this is typically drgdp2).
    """
    xl = pd.ExcelFile(path)
    sheet_candidates = [s for s in xl.sheet_names if str(s).strip().upper() == "RGDP"]
    if not sheet_candidates:
        sheet_candidates = [s for s in xl.sheet_names if "RGDP" in str(s).upper()]
    if not sheet_candidates:
        raise ValueError(f"Could not find an RGDP sheet in workbook: {path}")

    sheet = sheet_candidates[0]
    df = pd.read_excel(path, sheet_name=sheet)
    df.columns = [normalize_column_name(c).lower() for c in df.columns]

    year_col = next((c for c in df.columns if c == "year"), None)
    quarter_col = next((c for c in df.columns if c == "quarter"), None)
    if year_col is None or quarter_col is None:
        raise ValueError(f"RGDP sheet in {path} does not expose YEAR / QUARTER columns.")

    growth_cols = [c for c in df.columns if re.match(r"^d?rgdp\d+$", c)]
    if not growth_cols:
        raise ValueError(f"RGDP sheet in {path} does not expose drgdp* growth columns.")

    def _growth_suffix(col: str) -> int:
        match = re.search(r"(\d+)$", col)
        return int(match.group(1)) if match else 10**6

    current_horizon_col = sorted(growth_cols, key=_growth_suffix)[0]

    out = pd.DataFrame()
    out["quarter"] = [
        pd.Period(f"{int(y):04d}Q{int(q)}", freq="Q")
        if pd.notna(y) and pd.notna(q) and int(q) in {1, 2, 3, 4}
        else pd.NaT
        for y, q in zip(df[year_col], df[quarter_col])
    ]
    out[value_label] = pd.to_numeric(df[current_horizon_col], errors="coerce")
    out["source_sheet"] = sheet
    out["horizon_col"] = current_horizon_col
    out = out.dropna(subset=["quarter"]).reset_index(drop=True)
    return out


def load_simple_quarter_value_table(path: Path, value_label: str) -> pd.DataFrame:
    """
    Parse quarter/value workbooks with file-specific logic where needed.
    """
    name = path.name.lower()
    if "gdpplus" in name:
        return load_gdpplus_latest_table(path, value_label=value_label)

    if "meangrowth" in name or "mediangrowth" in name:
        return load_spf_growth_benchmark(path, value_label=value_label)

    best: Optional[pd.DataFrame] = None
    best_period_col: Optional[str] = None
    best_score = -1.0

    for _, _, df in _candidate_excel_tables(path):
        period_col = _find_period_column(df)
        if period_col is None:
            continue
        num_cols = [c for c in df.columns if c != period_col and pd.to_numeric(df[c], errors="coerce").notna().mean() > 0.5]
        if not num_cols:
            continue
        score = float(len(num_cols))
        if score > best_score:
            best = df[[period_col] + num_cols].copy()
            best_period_col = period_col
            best_score = score

    if best is None or best_period_col is None:
        raise ValueError(f"Could not parse quarter/value table from workbook: {path}")

    num_cols = [c for c in best.columns if c != best_period_col]
    value_col = num_cols[0]
    out = pd.DataFrame(
        {
            "quarter": best[best_period_col].map(lambda x: parse_periodish(x, freq_hint="Q")),
            value_label: pd.to_numeric(best[value_col], errors="coerce"),
        }
    )
    return out.dropna(subset=["quarter"]).reset_index(drop=True)

# --------------------------------------------------------------------------------------
# Target / truth construction
# --------------------------------------------------------------------------------------

def select_target_workbooks(repo_root: Path) -> Dict[str, Path]:
    raw_dir = repo_root / "data" / "raw"
    mapping = {}
    for candidate in TARGET_FILE_PRIORITY:
        p = raw_dir / candidate
        if p.exists():
            mapping[candidate] = p
    return mapping



def build_target_and_truth_objects(repo_root: Path) -> Dict[str, pd.DataFrame]:
    workbooks = select_target_workbooks(repo_root)
    if "routputMvQd.xlsx" not in workbooks and "ROUTPUTQvQd.xlsx" not in workbooks:
        raise FileNotFoundError("Could not find ROUTPUT workbook in data/raw.")

    target_vintage: Optional[pd.DataFrame] = None
    last_target_error: Optional[Exception] = None
    for candidate_name in ["routputMvQd.xlsx", "ROUTPUTQvQd.xlsx"]:
        candidate_path = workbooks.get(candidate_name)
        if candidate_path is None:
            continue
        try:
            target_vintage = load_routput_vintage_history(candidate_path)
            break
        except Exception as err:
            last_target_error = err

    if target_vintage is None:
        if last_target_error is not None:
            raise last_target_error
        raise ValueError("Could not parse any ROUTPUT workbook in data/raw.")

    truth_objects: Dict[str, pd.DataFrame] = {
        "target_vintage_table": target_vintage
    }

    truth_latest = (
        target_vintage.dropna(subset=["gdp_growth_annualized"])
        .sort_values(["obs_period", "vintage_period"])
        .groupby("obs_period", as_index=False)
        .tail(1)
        .rename(columns={"obs_period": "quarter"})
        [["quarter", "vintage_period", "gdp_growth_annualized"]]
        .rename(columns={"gdp_growth_annualized": "latest_rtdsm"})
        .reset_index(drop=True)
    )
    truth_objects["truth_latest"] = truth_latest

    release_path = workbooks.get("routput_first_second_third.xlsx")
    if release_path is not None:
        release_truth = load_release_truth_table(release_path)
        truth_objects["truth_release_table"] = release_truth
        if "third_release" in release_truth.columns:
            truth_objects["truth_third_release"] = (
                release_truth[["quarter", "third_release"]].dropna().reset_index(drop=True)
            )

    gdpplus_path = workbooks.get("GDPplus_Vintages.xlsx")
    if gdpplus_path is not None:
        truth_objects["truth_gdpplus"] = load_gdpplus_latest_table(gdpplus_path, "gdpplus")

    mean_path = workbooks.get("meanGrowth.xlsx")
    if mean_path is not None:
        truth_objects["spf_mean"] = load_spf_growth_benchmark(mean_path, "spf_mean")

    median_path = workbooks.get("medianGrowth.xlsx")
    if median_path is not None:
        truth_objects["spf_median"] = load_spf_growth_benchmark(median_path, "spf_median")

    return truth_objects

def quarter_of_vintage(vintage_period: pd.Period) -> pd.Period:
    return vintage_period.asfreq("Q")


def within_quarter_origin(vintage_period: pd.Period) -> int:
    return ((int(vintage_period.month) - 1) % 3) + 1


def get_quarter_end_month(quarter: pd.Period) -> pd.Period:
    return quarter.asfreq("M", how="end")


def target_history_asof(target_vintage_table: pd.DataFrame, vintage_period: pd.Period) -> pd.DataFrame:
    out = target_vintage_table[target_vintage_table["vintage_period"] == vintage_period].copy()
    out = out.sort_values("obs_period").reset_index(drop=True)
    return out


# --------------------------------------------------------------------------------------
# Stable subset selection, block coverage, factor mapping
# --------------------------------------------------------------------------------------

def rename_with_crosswalk(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [MNEMONIC_CROSSWALK.get(str(c), str(c)) for c in out.columns]
    out = _coalesce_duplicate_columns(out)
    return out


def select_monthly_panel(
    transformed_snapshot: pd.DataFrame,
    panel_mode: str = "stable",
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    transformed_snapshot = rename_with_crosswalk(transformed_snapshot)
    meta = stable_subset_metadata()

    if panel_mode == "stable":
        keep = [c for c in meta["mnemonic"].tolist() if c in transformed_snapshot.columns]
        panel = transformed_snapshot[keep].copy()
        panel_meta = meta[meta["mnemonic"].isin(keep)].copy().reset_index(drop=True)
        return panel, panel_meta

    # "full" fallback: keep every numeric series; attach block metadata where available.
    panel = transformed_snapshot.copy()
    panel_meta = (
        pd.DataFrame({"mnemonic": panel.columns})
        .merge(meta, on="mnemonic", how="left")
        .sort_values("mnemonic")
        .reset_index(drop=True)
    )
    return panel, panel_meta


def build_factor_mapping(monthly_cols: Sequence[str], panel_meta: pd.DataFrame, quarterly_target_name: str) -> Tuple[Dict[str, List[str]], Dict[Tuple[str, ...], int]]:
    block_lookup = panel_meta.set_index("mnemonic")["factor_name"].to_dict()

    # Count available block factors among the current monthly columns.
    available_block_factors = [block_lookup.get(col) for col in monthly_cols]
    available_block_factors = [str(v) for v in available_block_factors if pd.notna(v) and str(v) != ""]
    block_counts = pd.Series(available_block_factors).value_counts().sort_values(ascending=False)

    # DynamicFactorMQ requires the number of latent factors to be no greater than
    # the number of monthly observed variables. Keep the largest-loading blocks first
    # if the current vintage has too few monthly series after filtering.
    max_block_factors = max(0, len(monthly_cols) - 1)  # one slot reserved for the global factor
    kept_block_factors = set(block_counts.index[:max_block_factors].tolist())

    factors: Dict[str, List[str]] = {}
    for col in monthly_cols:
        f = ["global"]
        block_factor = block_lookup.get(col)
        if pd.notna(block_factor) and str(block_factor) in kept_block_factors:
            f.append(str(block_factor))
        factors[col] = f
    factors[quarterly_target_name] = ["global"]

    factor_orders = {("global",): 1}
    for block_factor in sorted(kept_block_factors):
        factor_orders[(block_factor,)] = 1
    return factors, factor_orders


def compute_block_coverage(panel: pd.DataFrame, panel_meta: pd.DataFrame, target_quarter: pd.Period) -> pd.DataFrame:
    q_start = target_quarter.asfreq("M", how="start")
    q_end = target_quarter.asfreq("M", how="end")
    months = pd.period_range(q_start, q_end, freq="M")

    block_rows = []
    for block, g in panel_meta.dropna(subset=["block"]).groupby("block"):
        cols = [c for c in g["mnemonic"].tolist() if c in panel.columns]
        if not cols:
            continue
        sub = panel.reindex(months)[cols]
        observed_counts = sub.notna().sum(axis=0)
        coverage = float((observed_counts / 3.0).mean()) if len(observed_counts) else np.nan
        block_rows.append(
            {
                "block": block,
                "block_label": BLOCK_LABELS.get(block, block),
                "factor_name": BLOCK_TO_FACTOR.get(block, block),
                "n_series": len(cols),
                "coverage": coverage,
            }
        )
    return pd.DataFrame(block_rows).sort_values("block").reset_index(drop=True)


def build_observed_months_by_series(panel: pd.DataFrame, target_quarter: pd.Period) -> Dict[str, List[str]]:
    q_start = target_quarter.asfreq("M", how="start")
    q_end = target_quarter.asfreq("M", how="end")
    months = pd.period_range(q_start, q_end, freq="M")
    out: Dict[str, List[str]] = {}
    sub = panel.reindex(months)
    for col in sub.columns:
        obs = [str(p) for p in months[sub[col].notna().values]]
        out[str(col)] = obs
    return out



def as_model_index(df: pd.DataFrame) -> pd.DataFrame:
    """
    Canonicalize model inputs to monthly / quarterly PeriodIndex objects.

    DynamicFactorMQ works cleanly with PeriodIndex, and this preserves the
    repository's first-day-of-month / first-day-of-quarter semantics without
    inventing month-end timestamps.
    """
    out = df.copy()
    idx = out.index

    if isinstance(idx, pd.PeriodIndex):
        return out.sort_index()

    if isinstance(idx, pd.DatetimeIndex):
        freqstr = getattr(idx, "freqstr", None) or pd.infer_freq(idx)
        if freqstr is None:
            raise ValueError(
                "Model inputs must have an inferable monthly or quarterly calendar. "
                "Got a DatetimeIndex with no fixed/inferable frequency."
            )
        freqstr = str(freqstr).upper()
        if freqstr.startswith("M"):
            out.index = idx.to_period("M")
        elif freqstr.startswith("Q"):
            out.index = idx.to_period("Q")
        else:
            raise ValueError(
                f"Model inputs must be monthly or quarterly. Got inferred frequency {freqstr!r}."
            )
        return out.sort_index()

    inferred = infer_period_frequency_from_values(list(idx))
    if inferred in {"M", "Q"}:
        parsed = [parse_periodish(v, freq_hint=inferred) for v in idx]
        if all(v is not None for v in parsed):
            out.index = pd.PeriodIndex(parsed, freq=inferred)
            return out.sort_index()

    raise TypeError(
        f"Model inputs must be indexed by PeriodIndex or DatetimeIndex. Got {type(idx).__name__}."
    )


def align_quarterly_target_to_monthly_support(
    monthly_panel: pd.DataFrame,
    quarterly_target: pd.DataFrame,
) -> pd.DataFrame:
    """
    Restrict quarterly observations to quarters whose quarter-end month lies
    inside the observed monthly panel span.

    Why this is necessary:
    statsmodels.DynamicFactorMQ.construct_endog concatenates the monthly panel
    index first and then appends any quarterly-only months that are outside the
    monthly span. In this repository, the GDP vintage history begins in 1947Q1
    while FRED-MD begins in 1959-01, so keeping the full quarterly history
    creates a non-monotonic monthly index inside statsmodels. Once that happens,
    statsmodels discards the date index and falls back to a generated integer
    index, which then breaks out-of-sample `news()` impact-date resolution.

    The mixed-frequency model only needs quarterly observations whose quarter-end
    month is on the monthly model calendar. We therefore keep the repository's
    native first-of-period semantics, but clip the quarterly history to the
    monthly support before fitting / applying the model.
    """
    monthly_panel = as_model_index(monthly_panel).sort_index()
    quarterly_target = as_model_index(quarterly_target).sort_index()

    if not isinstance(monthly_panel.index, pd.PeriodIndex) or not str(monthly_panel.index.freqstr).startswith("M"):
        raise ValueError(
            "Monthly panel must have a monthly PeriodIndex before mixed-frequency alignment. "
            f"Got {type(monthly_panel.index).__name__} / {getattr(monthly_panel.index, 'freqstr', None)!r}."
        )
    if not isinstance(quarterly_target.index, pd.PeriodIndex) or not str(quarterly_target.index.freqstr).startswith("Q"):
        raise ValueError(
            "Quarterly target must have a quarterly PeriodIndex before mixed-frequency alignment. "
            f"Got {type(quarterly_target.index).__name__} / {getattr(quarterly_target.index, 'freqstr', None)!r}."
        )

    if len(monthly_panel.index) == 0 or len(quarterly_target.index) == 0:
        return quarterly_target.copy()

    monthly_start = monthly_panel.index.min()
    monthly_end = monthly_panel.index.max()
    quarter_end_months = quarterly_target.index.asfreq("M", how="end")

    keep = (quarter_end_months >= monthly_start) & (quarter_end_months <= monthly_end)
    aligned = quarterly_target.loc[keep].copy()

    if aligned.empty:
        raise ValueError(
            "Quarterly target has no observations whose quarter-end month lies within the monthly panel span. "
            f"monthly_min={monthly_start}, monthly_max={monthly_end}, "
            f"quarterly_min={quarterly_target.index.min()}, quarterly_max={quarterly_target.index.max()}."
        )

    return aligned


def prepare_mixed_frequency_model_inputs(
    monthly_panel: pd.DataFrame,
    quarterly_target: pd.DataFrame,
    quarterly_target_name: str = "gdp_growth",
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Canonicalize model inputs and repair the monthly/quarterly support mismatch
    that otherwise causes statsmodels to ignore the date index.
    """
    monthly_panel = as_model_index(monthly_panel).sort_index()
    quarterly_target = as_model_index(quarterly_target).sort_index()

    if quarterly_target_name in quarterly_target.columns:
        quarterly_target = quarterly_target[[quarterly_target_name]].copy()

    quarterly_target = align_quarterly_target_to_monthly_support(
        monthly_panel,
        quarterly_target,
    )

    return monthly_panel, quarterly_target


def _prediction_index_from_model(model) -> pd.Index:
    """
    Return the actual date-like index that statsmodels should be able to use for
    prediction / news computations.
    """
    base_index = getattr(model, "_index", None)
    if isinstance(base_index, (pd.PeriodIndex, pd.DatetimeIndex)) and not bool(
        getattr(model, "_index_generated", False)
    ):
        return base_index

    row_labels = getattr(getattr(model, "data", None), "row_labels", None)
    if isinstance(row_labels, (pd.PeriodIndex, pd.DatetimeIndex)):
        return row_labels

    raise TypeError(
        "Could not locate a date-like prediction index on the statsmodels model. "
        f"base_index_type={type(base_index).__name__}, row_labels_type={type(row_labels).__name__}."
    )


def model_prediction_index_audit(model, impact_date=None) -> pd.DataFrame:
    """
    Human-readable audit of the statsmodels internal prediction index and row labels.
    """
    base_index = getattr(model, "_index", None)
    row_labels = getattr(getattr(model, "data", None), "row_labels", None)

    out: Dict[str, Any] = {
        "model_index_type": type(base_index).__name__ if base_index is not None else None,
        "model_index_freq": getattr(base_index, "freqstr", None) if base_index is not None else None,
        "model_index_generated": bool(getattr(model, "_index_generated", False)),
        "model_index_monotonic": bool(getattr(base_index, "is_monotonic_increasing", False)) if base_index is not None else None,
        "model_index_min": base_index.min() if isinstance(base_index, pd.Index) and len(base_index) else None,
        "model_index_max": base_index.max() if isinstance(base_index, pd.Index) and len(base_index) else None,
        "row_labels_type": type(row_labels).__name__ if row_labels is not None else None,
        "row_labels_freq": getattr(row_labels, "freqstr", None) if row_labels is not None else None,
        "row_labels_monotonic": bool(getattr(row_labels, "is_monotonic_increasing", False)) if row_labels is not None else None,
        "row_labels_min": row_labels.min() if isinstance(row_labels, pd.Index) and len(row_labels) else None,
        "row_labels_max": row_labels.max() if isinstance(row_labels, pd.Index) and len(row_labels) else None,
        "impact_date": impact_date,
        "impact_date_supported": None,
        "prediction_index_start": None,
        "prediction_index_end": None,
    }

    if impact_date is not None:
        try:
            start, end, oos, pred_ix = model._get_prediction_index(impact_date, impact_date)
            out["impact_date_supported"] = True
            out["prediction_loc_start"] = start
            out["prediction_loc_end"] = end
            out["prediction_oos"] = oos
            out["prediction_index_start"] = pred_ix[0] if pred_ix is not None and len(pred_ix) else None
            out["prediction_index_end"] = pred_ix[-1] if pred_ix is not None and len(pred_ix) else None
        except Exception as exc:
            out["impact_date_supported"] = False
            out["prediction_error"] = repr(exc)

    return pd.DataFrame([out])


def _coerce_impact_date_for_model_index(index: pd.Index, target_quarter: pd.Period):
    """
    Map the quarter-end impact month to the actual label type used by the
    statsmodels comparison dataset.

    The model should think in monthly periods for the mixed-frequency news
    decomposition. We therefore keep `2026-03` as the period label when the
    model index is PeriodIndex, and only convert to timestamps when the model
    is actually indexed by timestamps.
    """
    impact_month = get_quarter_end_month(target_quarter)
    if isinstance(index, pd.PeriodIndex):
        return impact_month.asfreq(index.freq)
    if isinstance(index, pd.DatetimeIndex):
        freqstr = str(getattr(index, "freqstr", "") or "")
        if freqstr.startswith("M") and not freqstr.startswith("MS"):
            return impact_month.to_timestamp(how="end")
        return impact_month.to_timestamp(how="start")
    raise TypeError(
        "News comparison index must be a PeriodIndex or DatetimeIndex; "
        f"got {type(index).__name__}."
    )


def _assert_news_impact_date_supported(index_or_model, impact_date, *, vintage: pd.Period, target_quarter: pd.Period) -> None:
    """
    Fail fast with a clear message if statsmodels will not be able to locate or
    extend the requested impact date on the *actual* prediction index used by
    `news()`.
    """
    if hasattr(index_or_model, "_get_prediction_index"):
        model = index_or_model
        try:
            model._get_prediction_index(impact_date, impact_date)
            return
        except Exception as exc:
            base_index = getattr(model, "_index", None)
            row_labels = getattr(getattr(model, "data", None), "row_labels", None)
            raise ValueError(
                "News impact date is incompatible with the actual statsmodels model prediction index. "
                f"vintage={vintage}, target_quarter={target_quarter}, impact_date={impact_date}, "
                f"model_index_type={type(base_index).__name__}, model_index_freq={getattr(base_index, 'freqstr', None)}, "
                f"model_index_generated={getattr(model, '_index_generated', None)}, "
                f"model_index_min={base_index.min() if isinstance(base_index, pd.Index) and len(base_index) else None}, "
                f"model_index_max={base_index.max() if isinstance(base_index, pd.Index) and len(base_index) else None}, "
                f"row_labels_type={type(row_labels).__name__}, row_labels_freq={getattr(row_labels, 'freqstr', None)}, "
                f"row_labels_min={row_labels.min() if isinstance(row_labels, pd.Index) and len(row_labels) else None}, "
                f"row_labels_max={row_labels.max() if isinstance(row_labels, pd.Index) and len(row_labels) else None}."
            ) from exc

    index = index_or_model
    from statsmodels.tsa.base.tsa_model import get_index_loc

    try:
        get_index_loc(impact_date, index)
    except Exception as exc:
        min_idx = index.min() if len(index) else None
        max_idx = index.max() if len(index) else None
        freqstr = getattr(index, "freqstr", None)
        raise ValueError(
            "News impact date is incompatible with the comparison model index. "
            f"vintage={vintage}, target_quarter={target_quarter}, impact_date={impact_date}, "
            f"index_type={type(index).__name__}, freq={freqstr}, min_index={min_idx}, max_index={max_idx}."
        ) from exc


# --------------------------------------------------------------------------------------
# Dynamic factor model fitting, nowcast extraction, news decomposition
# --------------------------------------------------------------------------------------

def select_factor_order(
    monthly_panel: pd.DataFrame,
    quarterly_target: pd.DataFrame,
    factors: Dict[str, List[str]],
    candidate_orders: Sequence[int],
    idiosyncratic_ar1: bool,
    em_maxiter: int,
    em_tolerance: float,
) -> int:
    best_p = int(candidate_orders[0])
    best_bic = np.inf
    for p in candidate_orders:
        factor_orders = {key: int(p) for key in {tuple(v for v in ["global"]), *[tuple([x]) for x in sorted({f for v in factors.values() for f in v if f != "global"})]}}
        # The line above creates duplicate global + blocks. Rebuild explicitly.
        factor_orders = {("global",): int(p)}
        block_factor_names = sorted({f for v in factors.values() for f in v if f not in {"global"}})
        for name in block_factor_names:
            factor_orders[(name,)] = int(p)
        model = DynamicFactorMQ(
            monthly_panel,
            endog_quarterly=quarterly_target,
            factors=factors,
            factor_orders=factor_orders,
            idiosyncratic_ar1=idiosyncratic_ar1,
            standardize=True,
        )
        try:
            res = model.fit_em(maxiter=em_maxiter, tolerance=em_tolerance, disp=False)
            bic = float(getattr(res, "bic", np.inf))
            if np.isfinite(bic) and bic < best_bic:
                best_bic = bic
                best_p = int(p)
        except Exception as err:
            warnings.warn(f"Factor order {p} failed during IC selection: {err}")
            continue
    return best_p


def em_convergence_delta_from_llf(llf_path: Sequence[float]) -> float:
    """
    Replicate the relative EM convergence criterion used by
    ``statsmodels.DynamicFactorMQ.fit_em``:

        delta_t = 2 * |llf_t - llf_{t-1}| / (|llf_t| + |llf_{t-1}|)

    This is the criterion actually compared with the configured tolerance, so
    diagnostics should report it directly rather than using an unrelated
    absolute-difference heuristic.
    """
    arr = np.asarray(llf_path, dtype=float)
    if arr.size < 2 or not np.isfinite(arr[-1]) or not np.isfinite(arr[-2]):
        return np.nan if arr.size < 2 else np.inf
    denom = np.abs(arr[-1]) + np.abs(arr[-2])
    if not np.isfinite(denom) or denom == 0:
        return np.inf
    return float(2.0 * np.abs(arr[-1] - arr[-2]) / denom)


def _coerce_start_params_for_model(model, start_params: Optional[Sequence[float]]):
    if start_params is None:
        return None

    param_names = list(getattr(model, "param_names", []) or [])
    if isinstance(start_params, pd.Series):
        if param_names and not set(param_names).issubset(set(start_params.index)):
            return None
        arr = (
            pd.to_numeric(start_params.reindex(param_names), errors="coerce").to_numpy(dtype=float)
            if param_names
            else pd.to_numeric(start_params, errors="coerce").to_numpy(dtype=float)
        )
    else:
        arr = np.asarray(start_params, dtype=float).reshape(-1)

    expected_k = int(getattr(model, "k_params", arr.size))
    if arr.size != expected_k or not np.all(np.isfinite(arr)):
        return None
    return arr


def fit_dfm_single_vintage(
    monthly_panel: pd.DataFrame,
    quarterly_target: pd.DataFrame,
    panel_meta: pd.DataFrame,
    factor_order: int,
    idiosyncratic_ar1: bool = True,
    em_maxiter: int = 100,
    em_tolerance: float = 1e-6,
    quarterly_target_name: str = "gdp_growth",
    start_params: Optional[Sequence[float]] = None,
):
    factors, factor_orders = build_factor_mapping(monthly_panel.columns, panel_meta, quarterly_target_name)
    factor_orders = {k: factor_order for k in factor_orders}

    monthly_panel_model, quarterly_target_model = prepare_mixed_frequency_model_inputs(
        monthly_panel,
        quarterly_target[[quarterly_target_name]],
        quarterly_target_name=quarterly_target_name,
    )

    model = DynamicFactorMQ(
        monthly_panel_model,
        endog_quarterly=quarterly_target_model,
        factors=factors,
        factor_orders=factor_orders,
        idiosyncratic_ar1=idiosyncratic_ar1,
        standardize=True,
        init_t0=False,
    )
    start_params_model = _coerce_start_params_for_model(model, start_params)
    results = model.fit_em(
        start_params=start_params_model,
        transformed=True,
        maxiter=em_maxiter,
        tolerance=em_tolerance,
        disp=False,
        full_output=True,
    )
    return model, results


def extract_nowcast_from_results(
    results,
    vintage_period: pd.Period,
    target_quarter: pd.Period,
    quarterly_target_name: str = "gdp_growth",
) -> float:
    """
    Extract E(g_q | I_v, theta_hat_v), preserving quarter semantics.

    If the quarter-end month is in-sample, use the filtered conditional mean for
    that month. If not, use the required monthly forecast horizon.
    """
    q_end_month_period = get_quarter_end_month(target_quarter)
    sample_end_raw = results.model.data.row_labels[-1]
    sample_end_period = sample_end_raw if isinstance(sample_end_raw, pd.Period) else pd.Period(pd.Timestamp(sample_end_raw), freq="M")

    if q_end_month_period <= sample_end_period:
        pred = results.get_prediction(
            start=q_end_month_period,
            end=q_end_month_period,
            information_set="filtered",
        ).predicted_mean
        return float(pred.loc[q_end_month_period, quarterly_target_name])

    steps = q_end_month_period.ordinal - sample_end_period.ordinal
    forecast = results.get_forecast(steps=steps).predicted_mean
    return float(forecast.iloc[-1][quarterly_target_name])


def _factor_anchor_signs(results, panel_meta: pd.DataFrame) -> Dict[str, int]:
    params = results.params
    signs: Dict[str, int] = {"global": 1}
    anchor_rows = panel_meta[panel_meta["anchor"] == True]
    for _, row in anchor_rows.iterrows():
        factor_name = row["factor_name"]
        if pd.isna(factor_name):
            continue
        factor_name = str(factor_name)
        mnemonic = str(row["mnemonic"])
        key = f"loading.{factor_name}->{mnemonic}"
        if key in params.index:
            signs[factor_name] = 1 if float(params.loc[key]) >= 0 else -1
    # Global factor orientation: prefer INDPRO if available, otherwise first loading.
    global_key_candidates = [f"loading.global->{m}" for m in anchor_rows["mnemonic"].astype(str)]
    for key in global_key_candidates:
        if key in params.index:
            signs["global"] = 1 if float(params.loc[key]) >= 0 else -1
            break
    return signs


def oriented_factor_states(
    results,
    panel_meta: pd.DataFrame,
    kind: str = "smoothed",
    standardize: bool = True,
    center: bool = True,
    scale_floor: float = 1e-6,
) -> pd.DataFrame:
    """
    Orient latent factors using economically interpretable anchor loadings and
    then normalize the exported state scale.

    The sign of a state-space factor is unidentified, and the scale is only
    identified up to a corresponding loading rescaling. Exporting centered,
    unit-variance states therefore improves cross-vintage comparability without
    altering the fitted signal that the model uses internally.
    """
    factors_df = getattr(results.factors, kind).copy()
    signs = _factor_anchor_signs(results, panel_meta)
    for col in factors_df.columns:
        sign = signs.get(str(col), 1)
        series = pd.to_numeric(factors_df[col], errors="coerce") * sign
        if standardize:
            location = float(series.mean(skipna=True)) if center else 0.0
            scale = float(series.std(skipna=True, ddof=0))
            if not np.isfinite(scale) or scale < float(scale_floor):
                scale = 1.0
            series = (series - location) / scale if center else (series / scale)
        factors_df[col] = series
    return factors_df


def flatten_news_results(
    news_results,
    panel_meta: pd.DataFrame,
    impacted_variable: str = "gdp_growth",
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    meta_lookup = panel_meta.set_index("mnemonic")[["block", "block_label", "factor_name"]].to_dict("index")
    details = news_results.details_by_update.reset_index().copy()
    details.columns = [normalize_column_name(c) for c in details.columns]
    # Columns typically include:
    # update_date, updated_variable, observed, forecast_(prev), impact_date,
    # impacted_variable, news, weight, impact
    updated_var_col = next((c for c in details.columns if c == "updated_variable"), "updated_variable")
    impact_var_col = next((c for c in details.columns if c == "impacted_variable"), "impacted_variable")

    if impact_var_col in details.columns:
        details = details[details[impact_var_col].astype(str) == impacted_variable].copy()

    if updated_var_col in details.columns:
        details["block"] = details[updated_var_col].map(lambda x: meta_lookup.get(str(x), {}).get("block"))
        details["block_label"] = details[updated_var_col].map(lambda x: meta_lookup.get(str(x), {}).get("block_label"))
        details["factor_name"] = details[updated_var_col].map(lambda x: meta_lookup.get(str(x), {}).get("factor_name"))

    numeric_cols = ["observed", "forecast_(prev)", "news", "weight", "impact"]
    for col in numeric_cols:
        if col in details.columns:
            details[col] = pd.to_numeric(details[col], errors="coerce")

    block_news = (
        details.groupby(["impact_date", "block", "block_label", "factor_name"], dropna=False)["impact"]
        .sum(min_count=1)
        .reset_index()
        .rename(columns={"impact": "signed_block_news"})
    )
    if not block_news.empty:
        block_news["abs_block_news"] = block_news["signed_block_news"].abs()

    return details, block_news


def make_diagnostics_row(
    vintage_period: pd.Period,
    target_quarter: pd.Period,
    monthly_panel: pd.DataFrame,
    results,
    factor_order: int,
) -> Dict[str, Any]:
    llf_path = np.asarray(results.mle_retvals.get("llf", []), dtype=float)
    base_index = getattr(results.model, "_index", None)
    row_labels = getattr(getattr(results.model, "data", None), "row_labels", None)
    mle_settings = getattr(results, "mle_settings", {}) or {}
    em_iterations = int(results.mle_retvals.get("iter", np.nan))
    em_tolerance_used = float(mle_settings.get("tolerance", np.nan))
    em_maxiter_used = int(mle_settings.get("maxiter", np.nan))
    em_delta_last = em_convergence_delta_from_llf(llf_path)
    converged_flag = bool(
        (np.isfinite(em_delta_last) and np.isfinite(em_tolerance_used) and em_delta_last <= em_tolerance_used)
        or (np.isfinite(em_iterations) and np.isfinite(em_maxiter_used) and em_iterations < em_maxiter_used)
    )

    return {
        "vintage_period": vintage_period,
        "vintage_timestamp_start": vintage_period.to_timestamp(),
        "target_quarter": target_quarter,
        "within_quarter_origin": within_quarter_origin(vintage_period),
        "n_monthly_series": int(monthly_panel.shape[1]),
        "n_monthly_obs": int(monthly_panel.shape[0]),
        "factor_order": int(factor_order),
        "llf_final": float(results.llf),
        "em_iterations": em_iterations,
        "em_tolerance_used": em_tolerance_used,
        "em_maxiter_used": em_maxiter_used,
        "em_convergence_delta_last": em_delta_last,
        "llf_path_json": json.dumps(llf_path.tolist()),
        "converged_flag": converged_flag,
        "sample_end_period": monthly_panel.index.max(),
        "model_index_generated": bool(getattr(results.model, "_index_generated", False)),
        "model_index_type": type(base_index).__name__ if base_index is not None else None,
        "model_index_monotonic": bool(getattr(base_index, "is_monotonic_increasing", False)) if base_index is not None else None,
        "row_labels_type": type(row_labels).__name__ if row_labels is not None else None,
        "row_labels_monotonic": bool(getattr(row_labels, "is_monotonic_increasing", False)) if row_labels is not None else None,
    }


# --------------------------------------------------------------------------------------
# End-to-end Layer 1 loop
# --------------------------------------------------------------------------------------

def choose_vintage_schedule(
    md_manifest: pd.DataFrame,
    start_quarter: str = "2000Q1",
    vintage_limit: Optional[int] = None,
) -> List[pd.Period]:
    vintages = md_manifest["vintage_period"].dropna().tolist()
    start_q = pd.Period(start_quarter, freq="Q")
    vintages = [v for v in vintages if quarter_of_vintage(v) >= start_q]
    if vintage_limit is not None:
        vintages = vintages[: int(vintage_limit)]
    return vintages


def build_quarterly_target_series_for_vintage(
    target_vintage_table: pd.DataFrame,
    vintage_period: pd.Period,
    quarterly_target_name: str = "gdp_growth",
) -> pd.DataFrame:
    hist = target_history_asof(target_vintage_table, vintage_period)
    out = hist[["obs_period", "gdp_growth_annualized"]].copy()
    out = out.rename(columns={"obs_period": "quarter", "gdp_growth_annualized": quarterly_target_name})
    out = out.dropna(subset=["quarter"]).set_index("quarter").sort_index()
    # Preserve quarter semantics: PeriodIndex('Q'), not month-end timestamps.
    out.index = pd.PeriodIndex(out.index, freq="Q")
    return out


def serialize_protocol(config: ProtocolConfig, path: Path) -> None:
    payload = asdict(config)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, default=str)


def export_table(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.suffix.lower() == ".csv":
        df.to_csv(path, index=False)
    elif path.suffix.lower() == ".parquet":
        try:
            df.to_parquet(path, index=False)
        except Exception:
            fallback = path.with_suffix(".csv")
            warnings.warn(
                f"Parquet export unavailable for {path.name}; wrote CSV fallback to {fallback.name} instead."
            )
            df.to_csv(fallback, index=False)
    else:
        raise ValueError(f"Unsupported export path: {path}")



def run_layer1_dfm(config: ProtocolConfig) -> Dict[str, pd.DataFrame]:
    repo_root = Path(config.repo_root)
    output_dir = ensure_directory(Path(config.output_dir))

    catalog = build_repo_catalog(repo_root)
    md_manifest = choose_canonical_md_manifest(catalog)
    qd_manifest = choose_canonical_qd_manifest(catalog)

    target_objects = build_target_and_truth_objects(repo_root)
    target_vintage_table = target_objects["target_vintage_table"]

    vintage_schedule = choose_vintage_schedule(
        md_manifest,
        start_quarter=config.benchmark_start_quarter,
        vintage_limit=config.vintage_limit,
    )

    target_vintage_set = set(target_vintage_table["vintage_period"].dropna().tolist())
    vintage_schedule = [v for v in vintage_schedule if v in target_vintage_set]

    if len(vintage_schedule) == 0:
        raise ValueError("No eligible monthly vintages found for the selected benchmark window after capping to available target vintages.")

    nowcast_rows: List[Dict[str, Any]] = []
    state_rows: List[pd.DataFrame] = []
    news_series_rows: List[pd.DataFrame] = []
    news_block_rows: List[pd.DataFrame] = []
    coverage_rows: List[pd.DataFrame] = []
    diagnostics_rows: List[Dict[str, Any]] = []

    prev_results = None
    prev_panel_meta: Optional[pd.DataFrame] = None
    prev_nowcast = np.nan
    prev_vintage = None

    for vintage in vintage_schedule:
        if not isinstance(vintage, pd.Period) or not str(vintage.freqstr).startswith("M"):
            raise ValueError(f"Vintage schedule must be monthly Periods. Got {vintage!r}.")

        md_row = md_manifest.loc[md_manifest["vintage_period"] == vintage]
        if md_row.empty:
            continue
        md_path = repo_root / md_row.iloc[0]["path"]

        snapshot = load_fred_snapshot(md_path, freq_hint="M")
        transformed = apply_tcodes_to_snapshot(snapshot)
        monthly_panel, panel_meta = select_monthly_panel(transformed, panel_mode=config.panel_mode)

        enough_data = monthly_panel.notna().sum() >= int(config.min_monthly_obs)
        monthly_panel = monthly_panel.loc[:, enough_data].copy()
        panel_meta = panel_meta[panel_meta["mnemonic"].isin(monthly_panel.columns)].copy().reset_index(drop=True)

        quarterly_target = build_quarterly_target_series_for_vintage(target_vintage_table, vintage)
        target_quarter = quarter_of_vintage(vintage)

        if not isinstance(target_quarter, pd.Period) or not str(target_quarter.freqstr).startswith("Q"):
            raise ValueError(f"Target quarter must be quarterly Period. Got {target_quarter!r}.")

        if monthly_panel.shape[1] == 0 or monthly_panel.shape[0] == 0:
            diagnostics_rows.append(
                {
                    "vintage_period": vintage,
                    "vintage_timestamp_start": vintage.to_timestamp(),
                    "target_quarter": target_quarter,
                    "within_quarter_origin": within_quarter_origin(vintage),
                    "n_monthly_series": int(monthly_panel.shape[1]),
                    "n_monthly_obs": int(monthly_panel.shape[0]),
                    "factor_order": np.nan,
                    "llf_final": np.nan,
                    "em_iterations": np.nan,
                    "llf_path_json": json.dumps([]),
                    "converged_flag": False,
                    "sample_end_period": monthly_panel.index.max() if len(monthly_panel.index) else pd.NaT,
                    "skipped_reason": "no monthly series after filtering",
                }
            )
            continue

        monthly_panel = as_model_index(monthly_panel)
        quarterly_target = as_model_index(quarterly_target)
        monthly_panel, quarterly_target = prepare_mixed_frequency_model_inputs(
            monthly_panel,
            quarterly_target,
            quarterly_target_name="gdp_growth",
        )

        if not isinstance(monthly_panel.index, pd.PeriodIndex) or not str(monthly_panel.index.freqstr).startswith("M"):
            raise ValueError(f"Monthly panel must have a monthly PeriodIndex. Got {type(monthly_panel.index).__name__} / {getattr(monthly_panel.index, 'freqstr', None)!r}.")
        if not isinstance(quarterly_target.index, pd.PeriodIndex) or not str(quarterly_target.index.freqstr).startswith("Q"):
            raise ValueError(f"Quarterly target must have a quarterly PeriodIndex. Got {type(quarterly_target.index).__name__} / {getattr(quarterly_target.index, 'freqstr', None)!r}.")

        if config.select_factor_order_per_vintage:
            factors, _ = build_factor_mapping(monthly_panel.columns, panel_meta, quarterly_target_name="gdp_growth")
            factor_order = select_factor_order(
                monthly_panel=monthly_panel,
                quarterly_target=quarterly_target[["gdp_growth"]],
                factors=factors,
                candidate_orders=config.candidate_factor_orders,
                idiosyncratic_ar1=config.idiosyncratic_ar1,
                em_maxiter=max(20, min(50, config.em_maxiter)),
                em_tolerance=config.em_tolerance,
            )
        else:
            factor_order = int(config.fixed_factor_order or config.candidate_factor_orders[0])

        start_params = None
        if config.warm_start_from_previous_vintage and prev_results is not None:
            prev_monthly_names = list(prev_results.model.endog_names[: prev_results.model.k_endog_M])
            current_monthly_names = list(monthly_panel.columns)
            if prev_monthly_names == current_monthly_names:
                start_params = getattr(prev_results, "params", None)

        model, results = fit_dfm_single_vintage(
            monthly_panel=monthly_panel,
            quarterly_target=quarterly_target,
            panel_meta=panel_meta,
            factor_order=factor_order,
            idiosyncratic_ar1=config.idiosyncratic_ar1,
            em_maxiter=config.em_maxiter,
            em_tolerance=config.em_tolerance,
            quarterly_target_name="gdp_growth",
            start_params=start_params,
        )

        nowcast_value = extract_nowcast_from_results(
            results=results,
            vintage_period=vintage,
            target_quarter=target_quarter,
            quarterly_target_name="gdp_growth",
        )

        coverage = compute_block_coverage(monthly_panel, panel_meta, target_quarter)
        coverage["vintage_period"] = vintage
        coverage["target_quarter"] = target_quarter
        coverage_rows.append(coverage)

        observed_months = build_observed_months_by_series(monthly_panel, target_quarter)

        factors_smoothed = oriented_factor_states(
            results,
            panel_meta,
            kind="smoothed",
            standardize=config.state_standardize,
            center=config.state_center,
            scale_floor=config.state_scale_floor,
        )
        current_state = factors_smoothed.loc[[factors_smoothed.index.max()]].copy()
        current_state["vintage_period"] = vintage
        current_state["target_quarter"] = target_quarter
        current_state["state_kind"] = "current_smoothed"
        state_rows.append(current_state.reset_index().rename(columns={"index": "state_period"}))

        pure_news_nowcast_value = np.nan
        pure_news_revision = np.nan
        full_refit_revision = np.nan if pd.isna(prev_nowcast) else nowcast_value - prev_nowcast
        reestimation_effect = np.nan

        if prev_results is not None and prev_vintage is not None:
            prev_meta_for_export = prev_panel_meta if prev_panel_meta is not None else panel_meta
            prev_factors_smoothed = oriented_factor_states(
                prev_results,
                prev_meta_for_export,
                kind="smoothed",
                standardize=config.state_standardize,
                center=config.state_center,
                scale_floor=config.state_scale_floor,
            )
            prev_state = prev_factors_smoothed.loc[[prev_factors_smoothed.index.max()]].copy()
            prev_state["vintage_period"] = vintage
            prev_state["target_quarter"] = target_quarter
            prev_state["state_kind"] = "previous_vintage_smoothed"
            state_rows.append(prev_state.reset_index().rename(columns={"index": "state_period"}))

            prev_monthly_names = list(prev_results.model.endog_names[: prev_results.model.k_endog_M])
            comparison_monthly = monthly_panel.reindex(columns=prev_monthly_names)
            comparison_monthly_model, quarterly_target_model = prepare_mixed_frequency_model_inputs(
                comparison_monthly,
                quarterly_target[["gdp_growth"]],
                quarterly_target_name="gdp_growth",
            )

            comparison_results = prev_results.apply(
                comparison_monthly_model,
                endog_quarterly=quarterly_target_model,
                copy_initialization=True,
            )
            pure_news_nowcast_value = extract_nowcast_from_results(
                results=comparison_results,
                vintage_period=vintage,
                target_quarter=target_quarter,
                quarterly_target_name="gdp_growth",
            )
            pure_news_revision = np.nan if pd.isna(prev_nowcast) else pure_news_nowcast_value - prev_nowcast
            reestimation_effect = (
                np.nan
                if pd.isna(full_refit_revision) or pd.isna(pure_news_revision)
                else full_refit_revision - pure_news_revision
            )

            comparison_prediction_index = _prediction_index_from_model(comparison_results.model)

            impact_date = _coerce_impact_date_for_model_index(
                comparison_prediction_index,
                target_quarter,
            )
            _assert_news_impact_date_supported(
                comparison_results.model,
                impact_date,
                vintage=vintage,
                target_quarter=target_quarter,
            )

            news = prev_results.news(
                comparison_results,
                comparison_type="updated",
                impact_date=impact_date,
                impacted_variable="gdp_growth",
                original_scale=True,
            )
            news_series, news_blocks = flatten_news_results(news, prev_meta_for_export, impacted_variable="gdp_growth")
            news_series["vintage_period"] = vintage
            news_series["target_quarter"] = target_quarter
            news_blocks["vintage_period"] = vintage
            news_blocks["target_quarter"] = target_quarter
            news_series_rows.append(news_series)
            news_block_rows.append(news_blocks)
        else:
            news_series_rows.append(pd.DataFrame({"vintage_period": [vintage], "target_quarter": [target_quarter]}))
            news_block_rows.append(pd.DataFrame({"vintage_period": [vintage], "target_quarter": [target_quarter]}))

        diag = make_diagnostics_row(
            vintage_period=vintage,
            target_quarter=target_quarter,
            monthly_panel=monthly_panel,
            results=results,
            factor_order=factor_order,
        )
        diagnostics_rows.append(diag)

        nowcast_rows.append(
            {
                "vintage_period": vintage,
                "vintage_timestamp_start": vintage.to_timestamp(),
                "target_quarter": target_quarter,
                "within_quarter_origin": within_quarter_origin(vintage),
                "dfm_nowcast": nowcast_value,
                "dfm_nowcast_pure_news_fixed_params": pure_news_nowcast_value,
                "dfm_nowcast_revision_from_previous": pure_news_revision,
                "dfm_nowcast_revision_full_refit": full_refit_revision,
                "dfm_nowcast_revision_reestimation_effect": reestimation_effect,
                "monthly_snapshot_path": str(md_row.iloc[0]["path"]),
                "observed_months_json": json.dumps(observed_months),
                "n_monthly_series": int(monthly_panel.shape[1]),
                "model_endog_mean_json": json.dumps(results.model._endog_mean.to_dict(), default=str),
                "model_endog_std_json": json.dumps(results.model._endog_std.to_dict(), default=str),
            }
        )

        prev_results = results
        prev_panel_meta = panel_meta.copy()
        prev_nowcast = nowcast_value
        prev_vintage = vintage

    nowcasts_df = pd.DataFrame(nowcast_rows)
    if nowcasts_df.empty:
        nowcasts_df = pd.DataFrame(
            columns=[
                "vintage_period",
                "vintage_timestamp_start",
                "target_quarter",
                "within_quarter_origin",
                "dfm_nowcast",
                "dfm_nowcast_pure_news_fixed_params",
                "dfm_nowcast_revision_from_previous",
                "dfm_nowcast_revision_full_refit",
                "dfm_nowcast_revision_reestimation_effect",
                "monthly_snapshot_path",
                "observed_months_json",
                "n_monthly_series",
                "model_endog_mean_json",
                "model_endog_std_json",
            ]
        )
    states_df = pd.concat(state_rows, ignore_index=True) if state_rows else pd.DataFrame()
    news_series_df = pd.concat(news_series_rows, ignore_index=True) if news_series_rows else pd.DataFrame()
    news_blocks_df = pd.concat(news_block_rows, ignore_index=True) if news_block_rows else pd.DataFrame()
    coverage_df = pd.concat(coverage_rows, ignore_index=True) if coverage_rows else pd.DataFrame()
    diagnostics_df = pd.DataFrame(diagnostics_rows)

    if "truth_third_release" in target_objects and "target_quarter" in nowcasts_df.columns:
        nowcasts_df = nowcasts_df.merge(
            target_objects["truth_third_release"],
            left_on="target_quarter",
            right_on="quarter",
            how="left",
        ).drop(columns=["quarter"])
        nowcasts_df["dfm_residual_third_release"] = nowcasts_df["third_release"] - nowcasts_df["dfm_nowcast"]

    if "truth_latest" in target_objects and "target_quarter" in nowcasts_df.columns:
        nowcasts_df = nowcasts_df.merge(
            target_objects["truth_latest"][["quarter", "latest_rtdsm"]],
            left_on="target_quarter",
            right_on="quarter",
            how="left",
        ).drop(columns=["quarter"])
        nowcasts_df["dfm_residual_latest_rtdsm"] = nowcasts_df["latest_rtdsm"] - nowcasts_df["dfm_nowcast"]

    if "truth_gdpplus" in target_objects and "target_quarter" in nowcasts_df.columns:
        nowcasts_df = nowcasts_df.merge(
            target_objects["truth_gdpplus"][["quarter", "gdpplus"]],
            left_on="target_quarter",
            right_on="quarter",
            how="left",
        ).drop(columns=["quarter"])
        nowcasts_df["dfm_residual_gdpplus"] = nowcasts_df["gdpplus"] - nowcasts_df["dfm_nowcast"]

    nowcasts_df = nowcasts_df.sort_values(["vintage_period", "target_quarter"]).reset_index(drop=True)

    if config.export_same_tau_residual_lags and "dfm_residual_third_release" in nowcasts_df.columns and len(nowcasts_df):
        warnings.warn(
            "Release-safe same-tau residual lags are disabled by default. "
            "Only enable export_same_tau_residual_lags after implementing an explicit availability calendar."
        )

    serialize_protocol(config, output_dir / "layer1_protocol.json")
    export_table(nowcasts_df, output_dir / "dfm_nowcasts.csv")
    if not states_df.empty:
        export_table(states_df, output_dir / "dfm_states.parquet")
    if not news_series_df.empty:
        export_table(news_series_df, output_dir / "dfm_news_series.csv")
    if not news_blocks_df.empty:
        export_table(news_blocks_df, output_dir / "dfm_news_blocks.csv")
    if not coverage_df.empty:
        export_table(coverage_df, output_dir / "dfm_coverage.csv")
    export_table(diagnostics_df, output_dir / "dfm_diagnostics.csv")
    export_table(md_manifest, output_dir / "vintage_manifest_monthly.csv")
    export_table(qd_manifest, output_dir / "vintage_manifest_quarterly.csv")
    export_table(catalog, output_dir / "repository_catalog.csv")

    return {
        "catalog": catalog,
        "monthly_manifest": md_manifest,
        "quarterly_manifest": qd_manifest,
        "nowcasts": nowcasts_df,
        "states": states_df,
        "news_series": news_series_df,
        "news_blocks": news_blocks_df,
        "coverage": coverage_df,
        "diagnostics": diagnostics_df,
        **target_objects,
    }

# --------------------------------------------------------------------------------------
# Presentation helpers for the notebook
# --------------------------------------------------------------------------------------

def protocol_summary_frame(config: ProtocolConfig) -> pd.DataFrame:
    payload = asdict(config)
    return pd.DataFrame({"setting": list(payload.keys()), "value": list(payload.values())})


def completion_checklist_frame(output_dir: Path) -> pd.DataFrame:
    expected = [
        ("layer1_protocol.json", ["layer1_protocol.json"]),
        ("dfm_nowcasts.csv", ["dfm_nowcasts.csv"]),
        ("dfm_states.parquet|csv", ["dfm_states.parquet", "dfm_states.csv"]),
        ("dfm_news_series.csv", ["dfm_news_series.csv"]),
        ("dfm_news_blocks.csv", ["dfm_news_blocks.csv"]),
        ("dfm_coverage.csv", ["dfm_coverage.csv"]),
        ("dfm_diagnostics.csv", ["dfm_diagnostics.csv"]),
        ("vintage_manifest_monthly.csv", ["vintage_manifest_monthly.csv"]),
        ("vintage_manifest_quarterly.csv", ["vintage_manifest_quarterly.csv"]),
        ("repository_catalog.csv", ["repository_catalog.csv"]),
    ]
    rows = []
    for label, candidates in expected:
        existing = [output_dir / c for c in candidates if (output_dir / c).exists()]
        rows.append(
            {
                "artifact": label,
                "exists": len(existing) > 0,
                "resolved_path": existing[0].name if existing else None,
                "size_bytes": existing[0].stat().st_size if existing else np.nan,
            }
        )
    return pd.DataFrame(rows)
