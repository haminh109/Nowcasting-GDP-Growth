
"""
Helper utilities for Layer 1 of the hybrid DFM + ML nowcasting project.

The code is designed to be repository-aware, vintage-aware, and careful about
the repository's first-day-of-month / first-day-of-quarter timestamp
convention. It avoids artificial month-end conversion and instead canonicalizes
time semantics with pandas PeriodIndex objects.
"""

from __future__ import annotations

import json
import math
import re
import warnings
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Mapping, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from statsmodels.tsa.statespace.dynamic_factor_mq import DynamicFactorMQ


# --------------------------------------------------------------------------------------
# Stable subset metadata from the uploaded data dictionary.
# --------------------------------------------------------------------------------------

STABLE_SUBSET_METADATA: List[Dict[str, Any]] = [
    # Real activity & income
    {"mnemonic": "RPI", "block": "real_activity_income", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "W875RX1", "block": "real_activity_income", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "INDPRO", "block": "real_activity_income", "tcode": 5, "anchor": True,  "alias": None},
    {"mnemonic": "IPFINAL", "block": "real_activity_income", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "IPCONGD", "block": "real_activity_income", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "IPBUSEQ", "block": "real_activity_income", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "IPMANSICS", "block": "real_activity_income", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "CUMFNS", "block": "real_activity_income", "tcode": 2, "anchor": False, "alias": None},

    # Labor market
    {"mnemonic": "CLF16OV", "block": "labor_market", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "UNRATE", "block": "labor_market", "tcode": 2, "anchor": False, "alias": None},
    {"mnemonic": "UEMPMEAN", "block": "labor_market", "tcode": 2, "anchor": False, "alias": None},
    {"mnemonic": "CLAIMSx", "block": "labor_market", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "PAYEMS", "block": "labor_market", "tcode": 5, "anchor": True,  "alias": None},
    {"mnemonic": "MANEMP", "block": "labor_market", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "AWHMAN", "block": "labor_market", "tcode": 1, "anchor": False, "alias": None},
    {"mnemonic": "CES3000000008", "block": "labor_market", "tcode": 6, "anchor": False, "alias": None},

    # Housing & construction
    {"mnemonic": "HOUST", "block": "housing_construction", "tcode": 4, "anchor": True,  "alias": None},
    {"mnemonic": "PERMIT", "block": "housing_construction", "tcode": 4, "anchor": False, "alias": None},
    {"mnemonic": "HOUSTNE", "block": "housing_construction", "tcode": 4, "anchor": False, "alias": None},
    {"mnemonic": "HOUSTMW", "block": "housing_construction", "tcode": 4, "anchor": False, "alias": None},
    {"mnemonic": "HOUSTS", "block": "housing_construction", "tcode": 4, "anchor": False, "alias": None},
    {"mnemonic": "HOUSTW", "block": "housing_construction", "tcode": 4, "anchor": False, "alias": None},

    # Demand / orders / inventories
    {"mnemonic": "DPCERA3M086SBEA", "block": "demand_orders_inventories", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "CMRMTSPLx", "block": "demand_orders_inventories", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "RETAILx", "block": "demand_orders_inventories", "tcode": 5, "anchor": True,  "alias": None},
    {"mnemonic": "ACOGNO", "block": "demand_orders_inventories", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "AMDMNOx", "block": "demand_orders_inventories", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "ANDENOx", "block": "demand_orders_inventories", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "AMDMUOx", "block": "demand_orders_inventories", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "BUSINVx", "block": "demand_orders_inventories", "tcode": 5, "anchor": False, "alias": None},
    {"mnemonic": "ISRATIOx", "block": "demand_orders_inventories", "tcode": 2, "anchor": False, "alias": None},

    # Prices & inflation
    {"mnemonic": "CPIAUCSL", "block": "prices_inflation", "tcode": 6, "anchor": True,  "alias": None},
    {"mnemonic": "CPIULFSL", "block": "prices_inflation", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "CUSR0000SAS", "block": "prices_inflation", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "CUSR0000SAC", "block": "prices_inflation", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "PCEPI", "block": "prices_inflation", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "DDURRG3M086SBEA", "block": "prices_inflation", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "DSERRG3M086SBEA", "block": "prices_inflation", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "WPSFD49207", "block": "prices_inflation", "tcode": 6, "anchor": False, "alias": "PPIFGS"},
    {"mnemonic": "OILPRICEx", "block": "prices_inflation", "tcode": 6, "anchor": False, "alias": None},

    # Financial conditions
    {"mnemonic": "FEDFUNDS", "block": "financial_conditions", "tcode": 2, "anchor": True,  "alias": None},
    {"mnemonic": "TB3MS", "block": "financial_conditions", "tcode": 2, "anchor": False, "alias": None},
    {"mnemonic": "GS10", "block": "financial_conditions", "tcode": 2, "anchor": False, "alias": None},
    {"mnemonic": "BAA", "block": "financial_conditions", "tcode": 2, "anchor": False, "alias": None},
    {"mnemonic": "T10YFFM", "block": "financial_conditions", "tcode": 1, "anchor": False, "alias": None},
    {"mnemonic": "BAAFFM", "block": "financial_conditions", "tcode": 1, "anchor": False, "alias": None},
    {"mnemonic": "BUSLOANS", "block": "financial_conditions", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "REALLN", "block": "financial_conditions", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "NONREVSL", "block": "financial_conditions", "tcode": 6, "anchor": False, "alias": None},
    {"mnemonic": "TWEXAFEGSMTHx", "block": "financial_conditions", "tcode": 5, "anchor": False, "alias": "TWEXMMTH"},
]

BLOCK_LABELS: Dict[str, str] = {
    "real_activity_income": "Real activity & income",
    "labor_market": "Labor market",
    "housing_construction": "Housing & construction",
    "demand_orders_inventories": "Demand / orders / inventories",
    "prices_inflation": "Prices & inflation",
    "financial_conditions": "Financial conditions",
}

BLOCK_TO_FACTOR: Dict[str, str] = {
    "real_activity_income": "real_activity",
    "labor_market": "labor",
    "housing_construction": "housing",
    "demand_orders_inventories": "demand",
    "prices_inflation": "prices",
    "financial_conditions": "financial",
}

MNEMONIC_CROSSWALK: Dict[str, str] = {
    "PPIFGS": "WPSFD49207",
    "TWEXMMTH": "TWEXAFEGSMTHx",
}

TARGET_FILE_PRIORITY = [
    "routputMvQd.xlsx",
    "ROUTPUTQvQd.xlsx",
    "routput_first_second_third.xlsx",
    "GDPplus_Vintages.xlsx",
    "meanGrowth.xlsx",
    "medianGrowth.xlsx",
]


@dataclass
class ProtocolConfig:
    repo_root: str
    output_dir: str
    benchmark_start_quarter: str = "2000Q1"
    panel_mode: str = "stable"  # "stable" is the reproducible block-structured default.
    truth_main: str = "third_release"
    truth_robustness: Tuple[str, ...] = ("latest_rtdsm", "gdpplus")
    candidate_factor_orders: Tuple[int, ...] = (1, 2)
    fixed_factor_order: Optional[int] = 1
    select_factor_order_per_vintage: bool = False
    idiosyncratic_ar1: bool = True
    em_maxiter: int = 100
    em_tolerance: float = 1e-6
    vintage_limit: Optional[int] = None
    min_monthly_obs: int = 24
    force_refit: bool = True


# --------------------------------------------------------------------------------------
# Generic helpers
# --------------------------------------------------------------------------------------

def stable_subset_metadata() -> pd.DataFrame:
    out = pd.DataFrame(STABLE_SUBSET_METADATA)
    out["block_label"] = out["block"].map(BLOCK_LABELS)
    out["factor_name"] = out["block"].map(BLOCK_TO_FACTOR)
    return out


def normalize_column_name(value: Any) -> str:
    value = "" if value is None else str(value)
    value = value.strip()
    value = re.sub(r"\s+", "_", value)
    return value


def locate_repo_root(start: Optional[Path] = None) -> Path:
    """
    Search upwards for a repository root containing data/raw.
    """
    if start is None:
        start = Path.cwd()
    start = start.resolve()
    candidates = [start] + list(start.parents)
    for root in candidates:
        if (root / "data" / "raw").exists():
            return root
    raise FileNotFoundError(
        "Could not locate repository root containing data/raw. "
        "Place the notebook in the repository or set REPO_ROOT manually."
    )


def ensure_directory(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def to_month_period_from_filename(year: int, month: int) -> pd.Period:
    return pd.Period(f"{int(year):04d}-{int(month):02d}", freq="M")


def parse_vintage_from_filename(filename: str) -> Optional[pd.Period]:
    """
    Parse repository vintage markers without forcing month-end timestamps.
    Returns a monthly Period when possible.
    """
    name = Path(filename).name

    patterns = [
        r"(?P<y>\d{4})-(?P<m>\d{2})-(?P<d>\d{2})\.csv$",
        r"(?P<y>\d{4})-(?P<m>\d{2})(?:-(?:MD|QD))?\.csv$",
        r"FRED-QD_(?P<y>\d{4})m(?P<m>\d{2})\.csv$",
    ]
    for pattern in patterns:
        match = re.search(pattern, name, flags=re.IGNORECASE)
        if match:
            y = int(match.group("y"))
            m = int(match.group("m"))
            return to_month_period_from_filename(y, m)

    return None


def classify_repo_file(path: Path) -> str:
    parts = tuple(path.parts)
    name = path.name
    if "sample" in parts:
        return "sample_csv"
    if "FRED-MD-MONTHLY" in parts:
        return "fred_md_current"
    if "Historical-vintages-of-FRED-MD-2015-01-to-2024-12" in parts:
        return "fred_md_hist_2015_2024"
    if "Historical FRED-MD Vintages Final" in parts:
        return "fred_md_hist_legacy"
    if "FRED-QD-QUARTERLY" in parts:
        return "fred_qd_current"
    if "Historical vintages of FRED-QD 2018-05 to 2024-12" in parts:
        return "fred_qd_hist_2018_2024"
    if name.lower().endswith(".xlsx"):
        return "excel_workbook"
    return "other"


def build_repo_catalog(repo_root: Path) -> pd.DataFrame:
    """
    Recursively audit repository files. Uses filename-based vintage parsing and
    keeps canonical vintage markers as monthly Periods.
    """
    rows: List[Dict[str, Any]] = []
    for path in sorted(repo_root.rglob("*")):
        if path.is_dir():
            continue
        rel = path.relative_to(repo_root)
        vintage = parse_vintage_from_filename(path.name)
        rows.append(
            {
                "path": str(rel),
                "stem": path.stem,
                "suffix": path.suffix.lower(),
                "group": classify_repo_file(rel),
                "size_bytes": path.stat().st_size,
                "vintage_period": vintage,
                "vintage_timestamp_start": vintage.to_timestamp() if vintage is not None else pd.NaT,
                "name": path.name,
            }
        )
    catalog = pd.DataFrame(rows)
    if not catalog.empty:
        catalog["size_mb"] = catalog["size_bytes"] / 1024**2
    return catalog


def choose_canonical_md_manifest(catalog: pd.DataFrame) -> pd.DataFrame:
    if catalog.empty:
        return catalog.copy()

    md = catalog[catalog["group"].isin({"fred_md_current", "fred_md_hist_2015_2024", "fred_md_hist_legacy"})].copy()
    md = md[md["vintage_period"].notna()].copy()
    if md.empty:
        return md

    priority = {
        "fred_md_hist_legacy": 1,
        "fred_md_hist_2015_2024": 2,
        "fred_md_current": 3,
    }
    md["priority"] = md["group"].map(priority)
    md = md.sort_values(["vintage_period", "priority", "path"]).drop_duplicates("vintage_period", keep="last")
    return md.sort_values("vintage_period").reset_index(drop=True)


def choose_canonical_qd_manifest(catalog: pd.DataFrame) -> pd.DataFrame:
    if catalog.empty:
        return catalog.copy()

    qd = catalog[catalog["group"].isin({"fred_qd_current", "fred_qd_hist_2018_2024"})].copy()
    qd = qd[qd["vintage_period"].notna()].copy()
    if qd.empty:
        return qd

    priority = {
        "fred_qd_hist_2018_2024": 1,
        "fred_qd_current": 2,
    }
    qd["priority"] = qd["group"].map(priority)
    qd = qd.sort_values(["vintage_period", "priority", "path"]).drop_duplicates("vintage_period", keep="last")
    return qd.sort_values("vintage_period").reset_index(drop=True)


def summarize_manifest(manifest: pd.DataFrame, label: str) -> pd.DataFrame:
    if manifest.empty:
        return pd.DataFrame([{"label": label, "count": 0, "min_vintage": pd.NaT, "max_vintage": pd.NaT}])
    return pd.DataFrame(
        [
            {
                "label": label,
                "count": len(manifest),
                "min_vintage": manifest["vintage_period"].min(),
                "max_vintage": manifest["vintage_period"].max(),
                "min_path": manifest.loc[manifest["vintage_period"].idxmin(), "path"],
                "max_path": manifest.loc[manifest["vintage_period"].idxmax(), "path"],
            }
        ]
    )


# --------------------------------------------------------------------------------------
# Timestamp / period parsing that respects first-of-month / first-of-quarter storage.
# --------------------------------------------------------------------------------------

_QUARTER_PATTERNS = [
    re.compile(r"^\s*(?P<y>\d{4})\s*[:\-]?\s*[Qq](?P<q>[1-4])\s*$"),
    re.compile(r"^\s*(?P<y>\d{4})\s*[Qq](?P<q>[1-4])\s*$"),
]


def parse_periodish(value: Any, freq_hint: Optional[str] = None) -> Optional[pd.Period]:
    """
    Parse month-like or quarter-like values to a pandas Period without shifting
    to month-end. The function intentionally preserves first-of-period semantics.
    """
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, pd.Period):
        return value
    if isinstance(value, pd.Timestamp):
        if freq_hint == "Q":
            return value.to_period("Q")
        # Default to month semantics unless the caller explicitly requests quarterly
        # parsing. This avoids reclassifying first-of-quarter timestamps as quarter
        # markers when they actually represent month labels like 2000-01-01.
        return value.to_period("M")

    text = str(value).strip()
    if text == "":
        return None

    for pattern in _QUARTER_PATTERNS:
        match = pattern.match(text)
        if match:
            y = int(match.group("y"))
            q = int(match.group("q"))
            return pd.Period(f"{y}Q{q}", freq="Q")

    # Handle "YYYYmMM" or "YYYYMMM" style labels
    match = re.match(r"^\s*(?P<y>\d{4})[mM](?P<m>\d{2})\s*$", text)
    if match:
        return pd.Period(f"{int(match.group('y')):04d}-{int(match.group('m')):02d}", freq="M")

    # Try full timestamps or first-day period markers.
    try:
        ts = pd.to_datetime(text, errors="raise")
        if freq_hint == "Q":
            return ts.to_period("Q")
        if freq_hint == "M":
            return ts.to_period("M")
        if ts.day == 1 and ts.month in (1, 4, 7, 10):
            # Quarter-like if quarter starts only.
            return ts.to_period("Q")
        return ts.to_period("M")
    except Exception:
        return None


def parse_timestamp_series(values: pd.Series, freq_hint: Optional[str] = None) -> pd.Series:
    """
    Parse to Timestamp series while respecting repository first-day conventions.
    """
    periods = values.map(lambda x: parse_periodish(x, freq_hint=freq_hint))
    out = pd.Series(pd.NaT, index=values.index, dtype="datetime64[ns]")
    for idx, p in periods.items():
        if p is not None and not pd.isna(p):
            out.loc[idx] = p.to_timestamp()
    return out


def infer_period_frequency_from_values(values: Sequence[Any]) -> Optional[str]:
    """
    Infer whether a collection of labels is monthly or quarterly, while
    respecting the repository's first-day timestamp convention.

    The logic intentionally avoids forcing first-of-quarter timestamps like
    2025-01-01 to quarterly unless the *sequence* behaves like quarterly data.
    """
    explicit_quarter_hits = 0
    timestamps: List[pd.Timestamp] = []
    for value in values:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            continue
        text = str(value).strip()
        if text == "":
            continue
        if any(pattern.match(text) for pattern in _QUARTER_PATTERNS):
            explicit_quarter_hits += 1
            try:
                p = parse_periodish(text, freq_hint="Q")
                if p is not None:
                    timestamps.append(p.to_timestamp())
            except Exception:
                pass
            continue
        try:
            ts = pd.to_datetime(value, errors="raise")
            timestamps.append(pd.Timestamp(ts))
        except Exception:
            continue

    if not timestamps:
        return None

    if explicit_quarter_hits / len(timestamps) > 0.5:
        return "Q"

    ordinals = np.array([ts.year * 12 + ts.month for ts in timestamps], dtype=int)
    ordinals = np.unique(np.sort(ordinals))
    quarter_like_month_sets = ({1, 4, 7, 10}, {3, 6, 9, 12})
    if len(ordinals) >= 2:
        diffs = np.diff(ordinals)
        if len(diffs) > 0:
            month_set = set((ordinals - 1) % 12 + 1)
            if np.mean(diffs == 3) > 0.8 and any(month_set.issubset(s) for s in quarter_like_month_sets):
                return "Q"
            if np.mean(diffs == 1) > 0.8:
                return "M"

    months = {(o - 1) % 12 + 1 for o in ordinals}
    if any(months.issubset(s) for s in quarter_like_month_sets) and len(ordinals) >= 4:
        return "Q"
    return "M"


# --------------------------------------------------------------------------------------
# CSV inspection and FRED snapshot loading
# --------------------------------------------------------------------------------------

def detect_date_column(df: pd.DataFrame) -> str:
    preferred = ["sasdate", "date", "observation_date", "DATE", "SASDATE"]
    for col in df.columns:
        if str(col) in preferred:
            return col
    # otherwise use the column with the highest parse success among the first few columns
    best_col = df.columns[0]
    best_score = -1.0
    for col in list(df.columns[: min(5, len(df.columns))]):
        score = pd.Series(df[col]).map(lambda x: parse_periodish(x) is not None).mean()
        if score > best_score:
            best_score = float(score)
            best_col = col
    return best_col


def _coalesce_duplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    if not df.columns.duplicated().any():
        return df
    pieces = []
    for col in pd.Index(df.columns).unique():
        sub = df.loc[:, df.columns == col]
        if sub.shape[1] == 1:
            pieces.append(sub.iloc[:, 0].rename(col))
        else:
            combined = sub.bfill(axis=1).iloc[:, 0].rename(col)
            pieces.append(combined)
    out = pd.concat(pieces, axis=1)
    out = out.loc[:, ~out.columns.duplicated()]
    return out


def _apply_mnemonic_crosswalk_to_columns(columns: Sequence[Any]) -> List[str]:
    mapped = []
    for col in columns:
        text = str(col)
        mapped.append(MNEMONIC_CROSSWALK.get(text, text))
    return mapped


def _leading_metadata_row_info(df: pd.DataFrame, date_col: str) -> Dict[str, Any]:
    metadata_rows: List[int] = []
    tcode_row_idx: Optional[int] = None
    for i in range(min(5, len(df))):
        label = str(df.iloc[i][date_col]).strip().lower()
        numeric_share = (
            pd.to_numeric(df.iloc[i].drop(labels=[date_col]), errors="coerce")
            .between(1, 7, inclusive="both")
            .mean()
        )
        is_metadata = (
            label.startswith("transform")
            or label.startswith("tcode")
            or label.startswith("factor")
            or numeric_share > 0.6
        )
        if is_metadata:
            metadata_rows.append(i)
            if (label.startswith("transform") or label.startswith("tcode") or numeric_share > 0.6) and tcode_row_idx is None:
                tcode_row_idx = i
        else:
            break
    return {
        "metadata_rows": metadata_rows,
        "metadata_row_count": len(metadata_rows),
        "tcode_row_idx": tcode_row_idx,
        "has_tcode_row": tcode_row_idx is not None,
    }


def inspect_csv_schema(path: Path, n_preview_rows: int = 6) -> Dict[str, Any]:
    """
    Read a CSV using the most likely FRED-style schema and return a human-readable inspection.
    """
    df = pd.read_csv(path, low_memory=False)
    date_col = detect_date_column(df)
    meta = _leading_metadata_row_info(df, date_col)

    preview_df = df.head(n_preview_rows).copy()
    raw_date_series = df[date_col].iloc[meta["metadata_row_count"]:].copy()
    inferred_freq = infer_period_frequency_from_values(raw_date_series.head(50).tolist())
    parsed_ts = parse_timestamp_series(raw_date_series, freq_hint=inferred_freq)
    parsed_ts = parsed_ts.dropna()

    out = {
        "path": str(path),
        "n_rows": len(df),
        "n_cols": df.shape[1],
        "date_col": str(date_col),
        "has_tcode_row": bool(meta["has_tcode_row"]),
        "metadata_row_count": int(meta["metadata_row_count"]),
        "preview": preview_df,
        "min_timestamp": parsed_ts.min() if not parsed_ts.empty else pd.NaT,
        "max_timestamp": parsed_ts.max() if not parsed_ts.empty else pd.NaT,
        "inferred_freq": inferred_freq,
        "columns_preview": list(map(str, df.columns[: min(15, len(df.columns))])),
    }
    return out


def load_fred_snapshot(path: Path, freq_hint: Optional[str] = None) -> Dict[str, Any]:
    """
    Load a FRED-MD / FRED-QD style vintage snapshot.

    The function preserves the repository's first-day timestamp convention by
    creating a PeriodIndex (M or Q) from the raw date column instead of
    artificially shifting dates to month-end.
    """
    df = pd.read_csv(path, low_memory=False)
    df.columns = [normalize_column_name(c) for c in df.columns]
    date_col = detect_date_column(df)
    meta = _leading_metadata_row_info(df, date_col)

    tcodes = pd.Series(dtype="Int64")
    if meta["tcode_row_idx"] is not None:
        raw_tcodes = pd.to_numeric(df.iloc[meta["tcode_row_idx"]].drop(labels=[date_col]), errors="coerce")
        tcodes = raw_tcodes.astype("Int64")

    if meta["metadata_row_count"] > 0:
        df = df.iloc[meta["metadata_row_count"]:].copy()

    df.columns = [normalize_column_name(c) for c in _apply_mnemonic_crosswalk_to_columns(df.columns)]
    date_col = normalize_column_name(MNEMONIC_CROSSWALK.get(date_col, date_col))

    # If freq_hint is missing, infer from the raw date column after stripping metadata rows.
    inferred_freq = freq_hint or infer_period_frequency_from_values(df[date_col].head(100).tolist()) or "M"
    date_ts = parse_timestamp_series(df[date_col], freq_hint=inferred_freq)

    numeric = df.drop(columns=[date_col]).apply(pd.to_numeric, errors="coerce")
    numeric.columns = _apply_mnemonic_crosswalk_to_columns(numeric.columns)
    numeric = _coalesce_duplicate_columns(numeric)

    if inferred_freq.upper().startswith("Q"):
        index = date_ts.dt.to_period("Q")
    else:
        index = date_ts.dt.to_period("M")

    valid = ~pd.isna(index)
    numeric = numeric.loc[valid].copy()
    index = index[valid]
    date_ts = date_ts.loc[valid]

    numeric.index = index
    numeric.index.name = "period"

    tcodes.index = _apply_mnemonic_crosswalk_to_columns(tcodes.index)
    tcodes = tcodes[~pd.Index(tcodes.index).duplicated(keep="last")]

    return {
        "path": str(path),
        "date_col": date_col,
        "tcode_row_present": bool(meta["has_tcode_row"]),
        "metadata_row_count": int(meta["metadata_row_count"]),
        "tcodes": tcodes,
        "data": numeric.sort_index(),
        "raw_timestamps": pd.Series(date_ts.values, index=index, name="timestamp_start").sort_index(),
        "freq": inferred_freq.upper()[0],
    }


# --------------------------------------------------------------------------------------
# Transformation logic
# --------------------------------------------------------------------------------------

def apply_tcode(series: pd.Series, tcode: int) -> pd.Series:
    """
    Apply official FRED-MD / FRED-QD transformation code.
    """
    s = pd.to_numeric(series, errors="coerce").astype(float)

    if tcode == 1:
        out = s
    elif tcode == 2:
        out = s.diff()
    elif tcode == 3:
        out = s.diff().diff()
    elif tcode == 4:
        out = np.log(s.where(s > 0))
    elif tcode == 5:
        out = np.log(s.where(s > 0)).diff()
    elif tcode == 6:
        out = np.log(s.where(s > 0)).diff().diff()
    elif tcode == 7:
        out = (s / s.shift(1) - 1.0).diff()
    else:
        raise ValueError(f"Unsupported tcode {tcode!r}")
    out.name = series.name
    return out


def apply_tcodes_to_snapshot(snapshot: Dict[str, Any], tcode_override: Optional[Mapping[str, int]] = None) -> pd.DataFrame:
    data = snapshot["data"].copy()
    tcodes = snapshot["tcodes"].copy()
    if tcode_override is not None:
        for key, value in tcode_override.items():
            tcodes.loc[key] = int(value)

    transformed = {}
    for col in data.columns:
        tcode = tcodes.get(col, pd.NA)
        if pd.isna(tcode):
            transformed[col] = data[col]
        else:
            transformed[col] = apply_tcode(data[col], int(tcode))
    out = pd.DataFrame(transformed, index=data.index)
    return out


# --------------------------------------------------------------------------------------
# Excel inspection and flexible extraction
# --------------------------------------------------------------------------------------

def inspect_excel_workbook(path: Path, preview_rows: int = 6, preview_cols: int = 8) -> pd.DataFrame:
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: List[Dict[str, Any]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        preview: List[List[Any]] = []
        for row in ws.iter_rows(min_row=1, max_row=preview_rows, max_col=preview_cols, values_only=True):
            preview.append(list(row))
        rows.append(
            {
                "sheet_name": sheet_name,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "preview_top_left": preview,
            }
        )
    return pd.DataFrame(rows)


def _read_sheet_as_values(path: Path, sheet_name: str, max_rows: Optional[int] = None, max_cols: Optional[int] = None) -> List[List[Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    rows: List[List[Any]] = []
    kwargs: Dict[str, Any] = {"values_only": True}
    if max_rows is not None:
        kwargs["max_row"] = max_rows
    if max_cols is not None:
        kwargs["max_col"] = max_cols
    for row in ws.iter_rows(**kwargs):
        rows.append(list(row))
    return rows


def _trim_2d(values: List[List[Any]]) -> List[List[Any]]:
    if not values:
        return values
    n_cols = max(len(r) for r in values)
    arr = [list(r) + [None] * (n_cols - len(r)) for r in values]

    # Trim empty bottom rows
    while arr and all(v in (None, "") or (isinstance(v, float) and math.isnan(v)) for v in arr[-1]):
        arr.pop()

    if not arr:
        return arr

    # Trim empty right columns
    last_nonempty = -1
    for j in range(n_cols):
        if any(row[j] not in (None, "") and not (isinstance(row[j], float) and math.isnan(row[j])) for row in arr):
            last_nonempty = j
    if last_nonempty >= 0:
        arr = [row[: last_nonempty + 1] for row in arr]
    return arr


def _score_matrix_candidate(values: List[List[Any]], header_row: int, index_col: int) -> Dict[str, Any]:
    trimmed = _trim_2d(values)
    if not trimmed or header_row >= len(trimmed):
        return {"score": -1}

    n_cols = max(len(r) for r in trimmed)
    if index_col >= n_cols:
        return {"score": -1}

    row_labels = trimmed[header_row][index_col + 1 :]
    col_labels = [trimmed[i][index_col] if index_col < len(trimmed[i]) else None for i in range(header_row + 1, len(trimmed))]

    parsed_rows = [parse_periodish(v) for v in row_labels]
    parsed_cols = [parse_periodish(v) for v in col_labels]

    row_count = sum(v is not None for v in parsed_rows)
    col_count = sum(v is not None for v in parsed_cols)

    if row_count < 3 or col_count < 3:
        return {"score": -1}

    body = []
    for i in range(header_row + 1, len(trimmed)):
        row = trimmed[i]
        numeric_row = row[index_col + 1 :]
        body.extend([pd.to_numeric(x, errors="coerce") for x in numeric_row if x is not None])
    body = pd.Series(body)
    numeric_share = body.notna().mean() if len(body) else 0.0

    row_freq = infer_period_frequency_from_values(row_labels)
    col_freq = infer_period_frequency_from_values(col_labels)

    score = row_count + col_count + 10.0 * numeric_share
    return {
        "score": score,
        "header_row": header_row,
        "index_col": index_col,
        "row_freq": row_freq,
        "col_freq": col_freq,
        "row_labels_parsed": parsed_rows,
        "col_labels_parsed": parsed_cols,
    }


def extract_best_period_matrix(path: Path, required_col_freq: Optional[str] = None) -> Dict[str, Any]:
    """
    Search workbook sheets for a vintage-by-observation matrix where either rows
    or columns look like monthly/quarterly periods.

    This is used for RTDSM-style ROUTPUT and GDPplus workbooks.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    candidates: List[Dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        values = _read_sheet_as_values(path, sheet_name)
        values = _trim_2d(values)
        if not values:
            continue

        n_rows = len(values)
        n_cols = max(len(r) for r in values)
        for header_row in range(min(6, n_rows)):
            for index_col in range(min(4, n_cols)):
                cand = _score_matrix_candidate(values, header_row, index_col)
                if cand["score"] < 0:
                    continue
                cand["sheet_name"] = sheet_name
                cand["values"] = values
                if required_col_freq is not None and cand.get("row_freq") != required_col_freq and cand.get("col_freq") != required_col_freq:
                    continue
                candidates.append(cand)

    if not candidates:
        raise ValueError(f"Could not find a date-like matrix in workbook: {path}")

    # Prefer matrices with quarterly observation labels somewhere.
    def ranking_key(c: Dict[str, Any]) -> Tuple[int, float]:
        has_required = int(c.get("row_freq") == required_col_freq or c.get("col_freq") == required_col_freq)
        return (has_required, float(c["score"]))

    best = sorted(candidates, key=ranking_key, reverse=True)[0]
    values = best["values"]
    header_row = int(best["header_row"])
    index_col = int(best["index_col"])

    raw_col_labels = values[header_row][index_col + 1 :]
    raw_row_labels = [values[i][index_col] if index_col < len(values[i]) else None for i in range(header_row + 1, len(values))]
    row_freq = infer_period_frequency_from_values(raw_row_labels)
    col_freq = infer_period_frequency_from_values(raw_col_labels)
    row_periods = [parse_periodish(v, freq_hint=row_freq) for v in raw_row_labels]
    col_periods = [parse_periodish(v, freq_hint=col_freq) for v in raw_col_labels]

    body = []
    for i in range(header_row + 1, len(values)):
        row = values[i] + [None] * (len(raw_col_labels) + index_col + 1 - len(values[i]))
        body.append([pd.to_numeric(x, errors="coerce") for x in row[index_col + 1 : index_col + 1 + len(raw_col_labels)]])
    matrix = pd.DataFrame(body, index=row_periods, columns=col_periods)
    matrix = matrix.dropna(how="all").dropna(axis=1, how="all")

    return {
        "sheet_name": best["sheet_name"],
        "header_row": header_row,
        "index_col": index_col,
        "row_freq": row_freq,
        "col_freq": col_freq,
        "matrix": matrix,
        "raw_row_labels": raw_row_labels,
        "raw_col_labels": raw_col_labels,
    }


def melt_vintage_matrix(matrix_info: Dict[str, Any], value_name: str = "value") -> pd.DataFrame:
    matrix = matrix_info["matrix"].copy()
    row_freq = matrix_info["row_freq"]
    col_freq = matrix_info["col_freq"]

    if row_freq not in {"M", "Q"} and col_freq not in {"M", "Q"}:
        raise ValueError("Matrix does not expose monthly/quarterly period axes.")

    if row_freq in {"M", "Q"} and col_freq == "Q":
        long = (
            matrix.stack(dropna=False)
            .rename(value_name)
            .rename_axis(index=["vintage_period", "obs_period"])
            .reset_index()
        )
        long["vintage_freq"] = row_freq
        long["obs_freq"] = col_freq
        return long

    if col_freq in {"M", "Q"} and row_freq == "Q":
        long = (
            matrix.T.stack(dropna=False)
            .rename(value_name)
            .rename_axis(index=["vintage_period", "obs_period"])
            .reset_index()
        )
        long["vintage_freq"] = col_freq
        long["obs_freq"] = row_freq
        return long

    raise ValueError("Could not orient matrix as vintage x observed-period.")


def load_routput_vintage_history(path: Path) -> pd.DataFrame:
    matrix_info = extract_best_period_matrix(path, required_col_freq="Q")
    long = melt_vintage_matrix(matrix_info, value_name="level")
    long = long[long["obs_freq"] == "Q"].copy()
    long = long.dropna(subset=["vintage_period", "obs_period"])
    long["vintage_timestamp_start"] = long["vintage_period"].map(lambda p: p.to_timestamp() if isinstance(p, pd.Period) else pd.NaT)
    long["quarter_timestamp_start"] = long["obs_period"].map(lambda p: p.to_timestamp() if isinstance(p, pd.Period) else pd.NaT)

    # Growth computed from levels within vintage.
    long = long.sort_values(["vintage_period", "obs_period"]).reset_index(drop=True)
    long["gdp_growth_annualized"] = (
        long.groupby("vintage_period")["level"]
        .transform(lambda s: 400.0 * np.log(s / s.shift(1)))
    )
    return long


def _candidate_excel_tables(path: Path) -> Iterator[Tuple[str, int, pd.DataFrame]]:
    xl = pd.ExcelFile(path)
    for sheet in xl.sheet_names:
        for header in range(0, 5):
            try:
                df = pd.read_excel(path, sheet_name=sheet, header=header)
            except Exception:
                continue
            if df is None or df.empty:
                continue
            df = df.copy()
            df.columns = [normalize_column_name(c).lower() for c in df.columns]
            yield sheet, header, df


def _find_period_column(df: pd.DataFrame) -> Optional[str]:
    best_col = None
    best_score = 0.0
    for col in df.columns:
        score = df[col].map(lambda x: parse_periodish(x, freq_hint="Q") is not None).mean()
        if score > best_score:
            best_score = float(score)
            best_col = col
    return best_col if best_score > 0.5 else None


def load_release_truth_table(path: Path) -> pd.DataFrame:
    """
    Attempt to load a first/second/third release GDP workbook into a tidy table.

    Expected output columns:
        quarter, first_release, second_release, third_release
    """
    best: Optional[pd.DataFrame] = None
    best_score = -1.0

    for sheet, header, df in _candidate_excel_tables(path):
        period_col = _find_period_column(df)
        if period_col is None:
            continue

        cols_lower = list(df.columns)
        score = 0.0
        release_cols = {}
        for key in ["first", "second", "third", "latest", "mean", "median", "gdpplus"]:
            matches = [c for c in cols_lower if key in str(c).lower()]
            if matches:
                release_cols[key] = matches[0]
                score += 1.0

        if score > best_score:
            best = df
            best_score = score

    if best is None:
        raise ValueError(f"Could not detect a release truth table in workbook: {path}")

    period_col = _find_period_column(best)
    assert period_col is not None
    out = pd.DataFrame()
    out["quarter"] = best[period_col].map(lambda x: parse_periodish(x, freq_hint="Q"))
    for candidate_name, canonical_name in [
        ("first", "first_release"),
        ("second", "second_release"),
        ("third", "third_release"),
        ("latest", "latest"),
        ("mean", "mean"),
        ("median", "median"),
        ("gdpplus", "gdpplus"),
    ]:
        matches = [c for c in best.columns if candidate_name in str(c).lower()]
        if matches:
            out[canonical_name] = pd.to_numeric(best[matches[0]], errors="coerce")

    out = out.dropna(subset=["quarter"]).reset_index(drop=True)
    return out


def load_simple_quarter_value_table(path: Path, value_label: str) -> pd.DataFrame:
    """
    Parse a workbook that contains quarter + value information (e.g. GDPplus,
    meanGrowth, medianGrowth) with heuristic header detection.
    """
    best: Optional[pd.DataFrame] = None
    best_score = -1.0

    for _, _, df in _candidate_excel_tables(path):
        period_col = _find_period_column(df)
        if period_col is None:
            continue
        num_cols = [c for c in df.columns if c != period_col and pd.to_numeric(df[c], errors="coerce").notna().mean() > 0.5]
        score = len(num_cols)
        if score > best_score and num_cols:
            best = df[[period_col] + num_cols].copy()
            best_score = score

    if best is None:
        raise ValueError(f"Could not parse quarter/value table from workbook: {path}")

    period_col = _find_period_column(best)
    assert period_col is not None
    num_cols = [c for c in best.columns if c != period_col]
    value_col = num_cols[0]

    out = pd.DataFrame(
        {
            "quarter": best[period_col].map(lambda x: parse_periodish(x, freq_hint="Q")),
            value_label: pd.to_numeric(best[value_col], errors="coerce"),
        }
    )
    return out.dropna(subset=["quarter"]).reset_index(drop=True)


# --------------------------------------------------------------------------------------
# Target / truth construction
# --------------------------------------------------------------------------------------

def select_target_workbooks(repo_root: Path) -> Dict[str, Path]:
    raw_dir = repo_root / "data" / "raw"
    mapping = {}
    for candidate in TARGET_FILE_PRIORITY:
        p = raw_dir / candidate
        if p.exists():
            mapping[candidate] = p
    return mapping


def build_target_and_truth_objects(repo_root: Path) -> Dict[str, pd.DataFrame]:
    workbooks = select_target_workbooks(repo_root)
    if "routputMvQd.xlsx" not in workbooks and "ROUTPUTQvQd.xlsx" not in workbooks:
        raise FileNotFoundError("Could not find ROUTPUT workbook in data/raw.")

    routput_path = workbooks.get("routputMvQd.xlsx", workbooks.get("ROUTPUTQvQd.xlsx"))
    assert routput_path is not None

    target_vintage = load_routput_vintage_history(routput_path)

    truth_objects: Dict[str, pd.DataFrame] = {
        "target_vintage_table": target_vintage
    }

    # Latest RTDSM truth can always be constructed from vintage history.
    truth_latest = (
        target_vintage.dropna(subset=["gdp_growth_annualized"])
        .sort_values(["obs_period", "vintage_period"])
        .groupby("obs_period", as_index=False)
        .tail(1)
        .rename(columns={"obs_period": "quarter"})
        [["quarter", "vintage_period", "gdp_growth_annualized"]]
        .rename(columns={"gdp_growth_annualized": "latest_rtdsm"})
        .reset_index(drop=True)
    )
    truth_objects["truth_latest"] = truth_latest

    release_path = workbooks.get("routput_first_second_third.xlsx")
    if release_path is not None:
        release_truth = load_release_truth_table(release_path)
        truth_objects["truth_release_table"] = release_truth
        if "third_release" in release_truth.columns:
            truth_objects["truth_third_release"] = (
                release_truth[["quarter", "third_release"]].dropna().reset_index(drop=True)
            )

    gdpplus_path = workbooks.get("GDPplus_Vintages.xlsx")
    if gdpplus_path is not None:
        truth_objects["truth_gdpplus"] = load_simple_quarter_value_table(gdpplus_path, "gdpplus")

    mean_path = workbooks.get("meanGrowth.xlsx")
    if mean_path is not None:
        truth_objects["spf_mean"] = load_simple_quarter_value_table(mean_path, "spf_mean")

    median_path = workbooks.get("medianGrowth.xlsx")
    if median_path is not None:
        truth_objects["spf_median"] = load_simple_quarter_value_table(median_path, "spf_median")

    return truth_objects


def quarter_of_vintage(vintage_period: pd.Period) -> pd.Period:
    return vintage_period.asfreq("Q")


def within_quarter_origin(vintage_period: pd.Period) -> int:
    return ((int(vintage_period.month) - 1) % 3) + 1


def get_quarter_end_month(quarter: pd.Period) -> pd.Period:
    return quarter.asfreq("M", how="end")


def target_history_asof(target_vintage_table: pd.DataFrame, vintage_period: pd.Period) -> pd.DataFrame:
    out = target_vintage_table[target_vintage_table["vintage_period"] == vintage_period].copy()
    out = out.sort_values("obs_period").reset_index(drop=True)
    return out


# --------------------------------------------------------------------------------------
# Stable subset selection, block coverage, factor mapping
# --------------------------------------------------------------------------------------

def rename_with_crosswalk(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [MNEMONIC_CROSSWALK.get(str(c), str(c)) for c in out.columns]
    out = _coalesce_duplicate_columns(out)
    return out


def select_monthly_panel(
    transformed_snapshot: pd.DataFrame,
    panel_mode: str = "stable",
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    transformed_snapshot = rename_with_crosswalk(transformed_snapshot)
    meta = stable_subset_metadata()

    if panel_mode == "stable":
        keep = [c for c in meta["mnemonic"].tolist() if c in transformed_snapshot.columns]
        panel = transformed_snapshot[keep].copy()
        panel_meta = meta[meta["mnemonic"].isin(keep)].copy().reset_index(drop=True)
        return panel, panel_meta

    # "full" fallback: keep every numeric series; attach block metadata where available.
    panel = transformed_snapshot.copy()
    panel_meta = (
        pd.DataFrame({"mnemonic": panel.columns})
        .merge(meta, on="mnemonic", how="left")
        .sort_values("mnemonic")
        .reset_index(drop=True)
    )
    return panel, panel_meta


def build_factor_mapping(monthly_cols: Sequence[str], panel_meta: pd.DataFrame, quarterly_target_name: str) -> Tuple[Dict[str, List[str]], Dict[Tuple[str, ...], int]]:
    block_lookup = panel_meta.set_index("mnemonic")["factor_name"].to_dict()

    # Count available block factors among the current monthly columns.
    available_block_factors = [block_lookup.get(col) for col in monthly_cols]
    available_block_factors = [str(v) for v in available_block_factors if pd.notna(v) and str(v) != ""]
    block_counts = pd.Series(available_block_factors).value_counts().sort_values(ascending=False)

    # DynamicFactorMQ requires the number of latent factors to be no greater than
    # the number of monthly observed variables. Keep the largest-loading blocks first
    # if the current vintage has too few monthly series after filtering.
    max_block_factors = max(0, len(monthly_cols) - 1)  # one slot reserved for the global factor
    kept_block_factors = set(block_counts.index[:max_block_factors].tolist())

    factors: Dict[str, List[str]] = {}
    for col in monthly_cols:
        f = ["global"]
        block_factor = block_lookup.get(col)
        if pd.notna(block_factor) and str(block_factor) in kept_block_factors:
            f.append(str(block_factor))
        factors[col] = f
    factors[quarterly_target_name] = ["global"]

    factor_orders = {("global",): 1}
    for block_factor in sorted(kept_block_factors):
        factor_orders[(block_factor,)] = 1
    return factors, factor_orders


def compute_block_coverage(panel: pd.DataFrame, panel_meta: pd.DataFrame, target_quarter: pd.Period) -> pd.DataFrame:
    q_start = target_quarter.asfreq("M", how="start")
    q_end = target_quarter.asfreq("M", how="end")
    months = pd.period_range(q_start, q_end, freq="M")

    block_rows = []
    for block, g in panel_meta.dropna(subset=["block"]).groupby("block"):
        cols = [c for c in g["mnemonic"].tolist() if c in panel.columns]
        if not cols:
            continue
        sub = panel.reindex(months)[cols]
        observed_counts = sub.notna().sum(axis=0)
        coverage = float((observed_counts / 3.0).mean()) if len(observed_counts) else np.nan
        block_rows.append(
            {
                "block": block,
                "block_label": BLOCK_LABELS.get(block, block),
                "factor_name": BLOCK_TO_FACTOR.get(block, block),
                "n_series": len(cols),
                "coverage": coverage,
            }
        )
    return pd.DataFrame(block_rows).sort_values("block").reset_index(drop=True)


def build_observed_months_by_series(panel: pd.DataFrame, target_quarter: pd.Period) -> Dict[str, List[str]]:
    q_start = target_quarter.asfreq("M", how="start")
    q_end = target_quarter.asfreq("M", how="end")
    months = pd.period_range(q_start, q_end, freq="M")
    out: Dict[str, List[str]] = {}
    sub = panel.reindex(months)
    for col in sub.columns:
        obs = [str(p) for p in months[sub[col].notna().values]]
        out[str(col)] = obs
    return out


def as_model_index(df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert PeriodIndex inputs to first-day Timestamp indexes for statsmodels,
    while preserving first-of-period semantics and avoiding artificial
    month-end conversion.
    """
    out = df.copy()
    if isinstance(out.index, pd.PeriodIndex):
        out.index = out.index.to_timestamp()
    return out.sort_index()


# --------------------------------------------------------------------------------------
# Dynamic factor model fitting, nowcast extraction, news decomposition
# --------------------------------------------------------------------------------------

def select_factor_order(
    monthly_panel: pd.DataFrame,
    quarterly_target: pd.DataFrame,
    factors: Dict[str, List[str]],
    candidate_orders: Sequence[int],
    idiosyncratic_ar1: bool,
    em_maxiter: int,
    em_tolerance: float,
) -> int:
    best_p = int(candidate_orders[0])
    best_bic = np.inf
    for p in candidate_orders:
        factor_orders = {key: int(p) for key in {tuple(v for v in ["global"]), *[tuple([x]) for x in sorted({f for v in factors.values() for f in v if f != "global"})]}}
        # The line above creates duplicate global + blocks. Rebuild explicitly.
        factor_orders = {("global",): int(p)}
        block_factor_names = sorted({f for v in factors.values() for f in v if f not in {"global"}})
        for name in block_factor_names:
            factor_orders[(name,)] = int(p)
        model = DynamicFactorMQ(
            monthly_panel,
            endog_quarterly=quarterly_target,
            factors=factors,
            factor_orders=factor_orders,
            idiosyncratic_ar1=idiosyncratic_ar1,
            standardize=True,
        )
        try:
            res = model.fit_em(maxiter=em_maxiter, tolerance=em_tolerance, disp=False)
            bic = float(getattr(res, "bic", np.inf))
            if np.isfinite(bic) and bic < best_bic:
                best_bic = bic
                best_p = int(p)
        except Exception as err:
            warnings.warn(f"Factor order {p} failed during IC selection: {err}")
            continue
    return best_p


def fit_dfm_single_vintage(
    monthly_panel: pd.DataFrame,
    quarterly_target: pd.DataFrame,
    panel_meta: pd.DataFrame,
    factor_order: int,
    idiosyncratic_ar1: bool = True,
    em_maxiter: int = 100,
    em_tolerance: float = 1e-6,
    quarterly_target_name: str = "gdp_growth",
):
    factors, factor_orders = build_factor_mapping(monthly_panel.columns, panel_meta, quarterly_target_name)
    factor_orders = {k: factor_order for k in factor_orders}

    monthly_panel_model = as_model_index(monthly_panel)
    quarterly_target_model = as_model_index(quarterly_target[[quarterly_target_name]])

    model = DynamicFactorMQ(
        monthly_panel_model,
        endog_quarterly=quarterly_target_model,
        factors=factors,
        factor_orders=factor_orders,
        idiosyncratic_ar1=idiosyncratic_ar1,
        standardize=True,
        init_t0=False,
    )
    results = model.fit_em(
        maxiter=em_maxiter,
        tolerance=em_tolerance,
        disp=False,
        full_output=True,
    )
    return model, results


def extract_nowcast_from_results(
    results,
    vintage_period: pd.Period,
    target_quarter: pd.Period,
    quarterly_target_name: str = "gdp_growth",
) -> float:
    """
    Extract E(g_q | I_v, theta_hat_v), preserving quarter semantics.

    If the quarter-end month is in-sample, use the filtered conditional mean for
    that month. If not, use the required monthly forecast horizon.
    """
    q_end_month_period = get_quarter_end_month(target_quarter)
    sample_end_raw = results.model.data.row_labels[-1]
    sample_end_period = sample_end_raw if isinstance(sample_end_raw, pd.Period) else pd.Period(pd.Timestamp(sample_end_raw), freq="M")

    if q_end_month_period <= sample_end_period:
        pred = results.get_prediction(
            start=q_end_month_period,
            end=q_end_month_period,
            information_set="filtered",
        ).predicted_mean
        return float(pred.loc[q_end_month_period, quarterly_target_name])

    steps = q_end_month_period.ordinal - sample_end_period.ordinal
    forecast = results.get_forecast(steps=steps).predicted_mean
    return float(forecast.iloc[-1][quarterly_target_name])


def _factor_anchor_signs(results, panel_meta: pd.DataFrame) -> Dict[str, int]:
    params = results.params
    signs: Dict[str, int] = {"global": 1}
    anchor_rows = panel_meta[panel_meta["anchor"] == True]
    for _, row in anchor_rows.iterrows():
        factor_name = row["factor_name"]
        if pd.isna(factor_name):
            continue
        factor_name = str(factor_name)
        mnemonic = str(row["mnemonic"])
        key = f"loading.{factor_name}->{mnemonic}"
        if key in params.index:
            signs[factor_name] = 1 if float(params.loc[key]) >= 0 else -1
    # Global factor orientation: prefer INDPRO if available, otherwise first loading.
    global_key_candidates = [f"loading.global->{m}" for m in anchor_rows["mnemonic"].astype(str)]
    for key in global_key_candidates:
        if key in params.index:
            signs["global"] = 1 if float(params.loc[key]) >= 0 else -1
            break
    return signs


def oriented_factor_states(results, panel_meta: pd.DataFrame, kind: str = "smoothed") -> pd.DataFrame:
    factors_df = getattr(results.factors, kind).copy()
    signs = _factor_anchor_signs(results, panel_meta)
    for col in factors_df.columns:
        sign = signs.get(str(col), 1)
        factors_df[col] = factors_df[col] * sign
    return factors_df


def flatten_news_results(
    news_results,
    panel_meta: pd.DataFrame,
    impacted_variable: str = "gdp_growth",
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    meta_lookup = panel_meta.set_index("mnemonic")[["block", "block_label", "factor_name"]].to_dict("index")
    details = news_results.details_by_update.reset_index().copy()
    details.columns = [normalize_column_name(c) for c in details.columns]
    # Columns typically include:
    # update_date, updated_variable, observed, forecast_(prev), impact_date,
    # impacted_variable, news, weight, impact
    updated_var_col = next((c for c in details.columns if c == "updated_variable"), "updated_variable")
    impact_var_col = next((c for c in details.columns if c == "impacted_variable"), "impacted_variable")

    if impact_var_col in details.columns:
        details = details[details[impact_var_col].astype(str) == impacted_variable].copy()

    if updated_var_col in details.columns:
        details["block"] = details[updated_var_col].map(lambda x: meta_lookup.get(str(x), {}).get("block"))
        details["block_label"] = details[updated_var_col].map(lambda x: meta_lookup.get(str(x), {}).get("block_label"))
        details["factor_name"] = details[updated_var_col].map(lambda x: meta_lookup.get(str(x), {}).get("factor_name"))

    numeric_cols = ["observed", "forecast_(prev)", "news", "weight", "impact"]
    for col in numeric_cols:
        if col in details.columns:
            details[col] = pd.to_numeric(details[col], errors="coerce")

    block_news = (
        details.groupby(["impact_date", "block", "block_label", "factor_name"], dropna=False)["impact"]
        .sum(min_count=1)
        .reset_index()
        .rename(columns={"impact": "signed_block_news"})
    )
    if not block_news.empty:
        block_news["abs_block_news"] = block_news["signed_block_news"].abs()

    return details, block_news


def make_diagnostics_row(
    vintage_period: pd.Period,
    target_quarter: pd.Period,
    monthly_panel: pd.DataFrame,
    results,
    factor_order: int,
) -> Dict[str, Any]:
    llf_path = np.asarray(results.mle_retvals.get("llf", []), dtype=float)
    return {
        "vintage_period": vintage_period,
        "vintage_timestamp_start": vintage_period.to_timestamp(),
        "target_quarter": target_quarter,
        "within_quarter_origin": within_quarter_origin(vintage_period),
        "n_monthly_series": int(monthly_panel.shape[1]),
        "n_monthly_obs": int(monthly_panel.shape[0]),
        "factor_order": int(factor_order),
        "llf_final": float(results.llf),
        "em_iterations": int(results.mle_retvals.get("iter", np.nan)),
        "llf_path_json": json.dumps(llf_path.tolist()),
        "converged_flag": bool(len(llf_path) < 2 or abs(llf_path[-1] - llf_path[-2]) <= 1e-4),
        "sample_end_period": monthly_panel.index.max(),
    }


# --------------------------------------------------------------------------------------
# End-to-end Layer 1 loop
# --------------------------------------------------------------------------------------

def choose_vintage_schedule(
    md_manifest: pd.DataFrame,
    start_quarter: str = "2000Q1",
    vintage_limit: Optional[int] = None,
) -> List[pd.Period]:
    vintages = md_manifest["vintage_period"].dropna().tolist()
    start_q = pd.Period(start_quarter, freq="Q")
    vintages = [v for v in vintages if quarter_of_vintage(v) >= start_q]
    if vintage_limit is not None:
        vintages = vintages[: int(vintage_limit)]
    return vintages


def build_quarterly_target_series_for_vintage(
    target_vintage_table: pd.DataFrame,
    vintage_period: pd.Period,
    quarterly_target_name: str = "gdp_growth",
) -> pd.DataFrame:
    hist = target_history_asof(target_vintage_table, vintage_period)
    out = hist[["obs_period", "gdp_growth_annualized"]].copy()
    out = out.rename(columns={"obs_period": "quarter", "gdp_growth_annualized": quarterly_target_name})
    out = out.dropna(subset=["quarter"]).set_index("quarter").sort_index()
    # Preserve quarter semantics: PeriodIndex('Q'), not month-end timestamps.
    out.index = pd.PeriodIndex(out.index, freq="Q")
    return out


def serialize_protocol(config: ProtocolConfig, path: Path) -> None:
    payload = asdict(config)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, default=str)


def export_table(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.suffix.lower() == ".csv":
        df.to_csv(path, index=False)
    elif path.suffix.lower() == ".parquet":
        try:
            df.to_parquet(path, index=False)
        except Exception:
            fallback = path.with_suffix(".csv")
            warnings.warn(
                f"Parquet export unavailable for {path.name}; wrote CSV fallback to {fallback.name} instead."
            )
            df.to_csv(fallback, index=False)
    else:
        raise ValueError(f"Unsupported export path: {path}")


def run_layer1_dfm(config: ProtocolConfig) -> Dict[str, pd.DataFrame]:
    repo_root = Path(config.repo_root)
    output_dir = ensure_directory(Path(config.output_dir))

    catalog = build_repo_catalog(repo_root)
    md_manifest = choose_canonical_md_manifest(catalog)
    qd_manifest = choose_canonical_qd_manifest(catalog)

    target_objects = build_target_and_truth_objects(repo_root)
    target_vintage_table = target_objects["target_vintage_table"]

    vintage_schedule = choose_vintage_schedule(
        md_manifest,
        start_quarter=config.benchmark_start_quarter,
        vintage_limit=config.vintage_limit,
    )

    if len(vintage_schedule) == 0:
        raise ValueError("No eligible monthly vintages found for the selected benchmark window.")

    nowcast_rows: List[Dict[str, Any]] = []
    state_rows: List[pd.DataFrame] = []
    news_series_rows: List[pd.DataFrame] = []
    news_block_rows: List[pd.DataFrame] = []
    coverage_rows: List[pd.DataFrame] = []
    diagnostics_rows: List[Dict[str, Any]] = []

    prev_results = None
    prev_nowcast = np.nan
    prev_vintage = None

    for vintage in vintage_schedule:
        md_row = md_manifest.loc[md_manifest["vintage_period"] == vintage]
        if md_row.empty:
            continue
        md_path = repo_root / md_row.iloc[0]["path"]

        snapshot = load_fred_snapshot(md_path, freq_hint="M")
        transformed = apply_tcodes_to_snapshot(snapshot)
        monthly_panel, panel_meta = select_monthly_panel(transformed, panel_mode=config.panel_mode)

        # Drop all-missing columns and enforce a minimal data requirement.
        enough_data = monthly_panel.notna().sum() >= int(config.min_monthly_obs)
        monthly_panel = monthly_panel.loc[:, enough_data].copy()
        panel_meta = panel_meta[panel_meta["mnemonic"].isin(monthly_panel.columns)].copy().reset_index(drop=True)

        quarterly_target = build_quarterly_target_series_for_vintage(target_vintage_table, vintage)

        # Ensure benchmark quarter exists in current target history window.
        target_quarter = quarter_of_vintage(vintage)

        if monthly_panel.shape[1] == 0 or monthly_panel.shape[0] == 0:
            diagnostics_rows.append(
                {
                    "vintage_period": vintage,
                    "vintage_timestamp_start": vintage.to_timestamp(),
                    "target_quarter": target_quarter,
                    "within_quarter_origin": within_quarter_origin(vintage),
                    "n_monthly_series": int(monthly_panel.shape[1]),
                    "n_monthly_obs": int(monthly_panel.shape[0]),
                    "factor_order": np.nan,
                    "llf_final": np.nan,
                    "em_iterations": np.nan,
                    "llf_path_json": json.dumps([]),
                    "converged_flag": False,
                    "sample_end_period": monthly_panel.index.max() if len(monthly_panel.index) else pd.NaT,
                    "skipped_reason": "no monthly series after filtering",
                }
            )
            continue

        # Choose factor order.
        if config.select_factor_order_per_vintage:
            factors, _ = build_factor_mapping(monthly_panel.columns, panel_meta, quarterly_target_name="gdp_growth")
            factor_order = select_factor_order(
                monthly_panel=monthly_panel,
                quarterly_target=quarterly_target[["gdp_growth"]],
                factors=factors,
                candidate_orders=config.candidate_factor_orders,
                idiosyncratic_ar1=config.idiosyncratic_ar1,
                em_maxiter=max(20, min(50, config.em_maxiter)),
                em_tolerance=config.em_tolerance,
            )
        else:
            factor_order = int(config.fixed_factor_order or config.candidate_factor_orders[0])

        model, results = fit_dfm_single_vintage(
            monthly_panel=monthly_panel,
            quarterly_target=quarterly_target,
            panel_meta=panel_meta,
            factor_order=factor_order,
            idiosyncratic_ar1=config.idiosyncratic_ar1,
            em_maxiter=config.em_maxiter,
            em_tolerance=config.em_tolerance,
            quarterly_target_name="gdp_growth",
        )

        nowcast_value = extract_nowcast_from_results(
            results=results,
            vintage_period=vintage,
            target_quarter=target_quarter,
            quarterly_target_name="gdp_growth",
        )

        # Coverage and observed-month audit
        coverage = compute_block_coverage(monthly_panel, panel_meta, target_quarter)
        coverage["vintage_period"] = vintage
        coverage["target_quarter"] = target_quarter
        coverage_rows.append(coverage)

        observed_months = build_observed_months_by_series(monthly_panel, target_quarter)

        # Oriented current and lagged factor states
        factors_smoothed = oriented_factor_states(results, panel_meta, kind="smoothed")
        current_state = factors_smoothed.loc[[factors_smoothed.index.max()]].copy()
        current_state["vintage_period"] = vintage
        current_state["target_quarter"] = target_quarter
        current_state["state_kind"] = "current_smoothed"
        state_rows.append(current_state.reset_index().rename(columns={"index": "state_period"}))

        if prev_results is not None:
            prev_factors_smoothed = oriented_factor_states(prev_results, panel_meta, kind="smoothed")
            prev_state = prev_factors_smoothed.loc[[prev_factors_smoothed.index.max()]].copy()
            prev_state["vintage_period"] = vintage
            prev_state["target_quarter"] = target_quarter
            prev_state["state_kind"] = "previous_vintage_smoothed"
            state_rows.append(prev_state.reset_index().rename(columns={"index": "state_period"}))

            # News decomposition uses previous vintage parameters and updated current-vintage data.
            # Align the current monthly panel to the variable ordering used by the previous model;
            # newly absent series remain explicit NaN rather than being dropped.
            prev_monthly_names = list(prev_results.model.endog_names[: prev_results.model.k_endog_M])
            comparison_monthly = monthly_panel.reindex(columns=prev_monthly_names)
            news = prev_results.news(
                as_model_index(comparison_monthly),
                comparison_type="updated",
                impact_date=get_quarter_end_month(target_quarter),
                impacted_variable="gdp_growth",
                endog_quarterly=as_model_index(quarterly_target[["gdp_growth"]]),
                original_scale=True,
            )
            news_series, news_blocks = flatten_news_results(news, panel_meta, impacted_variable="gdp_growth")
            news_series["vintage_period"] = vintage
            news_series["target_quarter"] = target_quarter
            news_blocks["vintage_period"] = vintage
            news_blocks["target_quarter"] = target_quarter
            news_series_rows.append(news_series)
            news_block_rows.append(news_blocks)
        else:
            # Seed empty tables for the first vintage.
            news_series_rows.append(pd.DataFrame({"vintage_period": [vintage], "target_quarter": [target_quarter]}))
            news_block_rows.append(pd.DataFrame({"vintage_period": [vintage], "target_quarter": [target_quarter]}))

        diag = make_diagnostics_row(
            vintage_period=vintage,
            target_quarter=target_quarter,
            monthly_panel=monthly_panel,
            results=results,
            factor_order=factor_order,
        )
        diagnostics_rows.append(diag)

        nowcast_rows.append(
            {
                "vintage_period": vintage,
                "vintage_timestamp_start": vintage.to_timestamp(),
                "target_quarter": target_quarter,
                "within_quarter_origin": within_quarter_origin(vintage),
                "dfm_nowcast": nowcast_value,
                "dfm_nowcast_revision_from_previous": np.nan if pd.isna(prev_nowcast) else nowcast_value - prev_nowcast,
                "monthly_snapshot_path": str(md_row.iloc[0]["path"]),
                "observed_months_json": json.dumps(observed_months),
                "n_monthly_series": int(monthly_panel.shape[1]),
                "model_endog_mean_json": json.dumps(results.model._endog_mean.to_dict(), default=str),
                "model_endog_std_json": json.dumps(results.model._endog_std.to_dict(), default=str),
            }
        )

        prev_results = results
        prev_nowcast = nowcast_value
        prev_vintage = vintage

    nowcasts_df = pd.DataFrame(nowcast_rows)
    if nowcasts_df.empty:
        nowcasts_df = pd.DataFrame(
            columns=[
                "vintage_period",
                "vintage_timestamp_start",
                "target_quarter",
                "within_quarter_origin",
                "dfm_nowcast",
                "dfm_nowcast_revision_from_previous",
                "monthly_snapshot_path",
                "observed_months_json",
                "n_monthly_series",
                "model_endog_mean_json",
                "model_endog_std_json",
            ]
        )
    states_df = pd.concat(state_rows, ignore_index=True) if state_rows else pd.DataFrame()
    news_series_df = pd.concat(news_series_rows, ignore_index=True) if news_series_rows else pd.DataFrame()
    news_blocks_df = pd.concat(news_block_rows, ignore_index=True) if news_block_rows else pd.DataFrame()
    coverage_df = pd.concat(coverage_rows, ignore_index=True) if coverage_rows else pd.DataFrame()
    diagnostics_df = pd.DataFrame(diagnostics_rows)

    # Merge truths that are already available as Layer 1 support objects.
    if "truth_third_release" in target_objects and "target_quarter" in nowcasts_df.columns:
        nowcasts_df = nowcasts_df.merge(
            target_objects["truth_third_release"],
            left_on="target_quarter",
            right_on="quarter",
            how="left",
        ).drop(columns=["quarter"])
        nowcasts_df["dfm_residual_third_release"] = nowcasts_df["third_release"] - nowcasts_df["dfm_nowcast"]

    if "truth_latest" in target_objects and "target_quarter" in nowcasts_df.columns:
        nowcasts_df = nowcasts_df.merge(
            target_objects["truth_latest"][["quarter", "latest_rtdsm"]],
            left_on="target_quarter",
            right_on="quarter",
            how="left",
        ).drop(columns=["quarter"])
        nowcasts_df["dfm_residual_latest_rtdsm"] = nowcasts_df["latest_rtdsm"] - nowcasts_df["dfm_nowcast"]

    if "truth_gdpplus" in target_objects and "target_quarter" in nowcasts_df.columns:
        nowcasts_df = nowcasts_df.merge(
            target_objects["truth_gdpplus"],
            left_on="target_quarter",
            right_on="quarter",
            how="left",
        ).drop(columns=["quarter"])
        nowcasts_df["dfm_residual_gdpplus"] = nowcasts_df["gdpplus"] - nowcasts_df["dfm_nowcast"]

    # Layer 2 lag features, same within-quarter origin only when truth is available.
    if "dfm_residual_third_release" in nowcasts_df.columns and len(nowcasts_df):
        nowcasts_df = nowcasts_df.sort_values(["within_quarter_origin", "target_quarter", "vintage_period"]).reset_index(drop=True)
        nowcasts_df["residual_lag1_same_tau"] = nowcasts_df.groupby("within_quarter_origin")["dfm_residual_third_release"].shift(1)
        nowcasts_df["residual_lag2_same_tau"] = nowcasts_df.groupby("within_quarter_origin")["dfm_residual_third_release"].shift(2)

    # Export
    serialize_protocol(config, output_dir / "layer1_protocol.json")
    export_table(nowcasts_df, output_dir / "dfm_nowcasts.csv")
    if not states_df.empty:
        export_table(states_df, output_dir / "dfm_states.parquet")
    if not news_series_df.empty:
        export_table(news_series_df, output_dir / "dfm_news_series.csv")
    if not news_blocks_df.empty:
        export_table(news_blocks_df, output_dir / "dfm_news_blocks.csv")
    if not coverage_df.empty:
        export_table(coverage_df, output_dir / "dfm_coverage.csv")
    export_table(diagnostics_df, output_dir / "dfm_diagnostics.csv")
    export_table(md_manifest, output_dir / "vintage_manifest_monthly.csv")
    export_table(qd_manifest, output_dir / "vintage_manifest_quarterly.csv")
    export_table(catalog, output_dir / "repository_catalog.csv")

    return {
        "catalog": catalog,
        "monthly_manifest": md_manifest,
        "quarterly_manifest": qd_manifest,
        "nowcasts": nowcasts_df,
        "states": states_df,
        "news_series": news_series_df,
        "news_blocks": news_blocks_df,
        "coverage": coverage_df,
        "diagnostics": diagnostics_df,
        **target_objects,
    }


# --------------------------------------------------------------------------------------
# Presentation helpers for the notebook
# --------------------------------------------------------------------------------------

def protocol_summary_frame(config: ProtocolConfig) -> pd.DataFrame:
    payload = asdict(config)
    return pd.DataFrame({"setting": list(payload.keys()), "value": list(payload.values())})


def completion_checklist_frame(output_dir: Path) -> pd.DataFrame:
    expected = [
        ("layer1_protocol.json", ["layer1_protocol.json"]),
        ("dfm_nowcasts.csv", ["dfm_nowcasts.csv"]),
        ("dfm_states.parquet|csv", ["dfm_states.parquet", "dfm_states.csv"]),
        ("dfm_news_series.csv", ["dfm_news_series.csv"]),
        ("dfm_news_blocks.csv", ["dfm_news_blocks.csv"]),
        ("dfm_coverage.csv", ["dfm_coverage.csv"]),
        ("dfm_diagnostics.csv", ["dfm_diagnostics.csv"]),
        ("vintage_manifest_monthly.csv", ["vintage_manifest_monthly.csv"]),
        ("vintage_manifest_quarterly.csv", ["vintage_manifest_quarterly.csv"]),
        ("repository_catalog.csv", ["repository_catalog.csv"]),
    ]
    rows = []
    for label, candidates in expected:
        existing = [output_dir / c for c in candidates if (output_dir / c).exists()]
        rows.append(
            {
                "artifact": label,
                "exists": len(existing) > 0,
                "resolved_path": existing[0].name if existing else None,
                "size_bytes": existing[0].stat().st_size if existing else np.nan,
            }
        )
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------------------
# Repository-specific parser overrides (patched after direct inspection of RTDSM / GDPplus / SPF layouts)
# --------------------------------------------------------------------------------------

_RTDSM_NUMERIC_QUARTER_PATTERN = re.compile(r"^\s*(?P<y>\d{4})\s*:\s*(?P<q>0?[1-4])\s*$")
_RTDSM_MONTH_VINTAGE_PATTERN = re.compile(r"^\s*[A-Za-z_]+(?P<yy>\d{2})M(?P<m>\d{1,2})\s*$", re.IGNORECASE)
_RTDSM_QUARTER_VINTAGE_PATTERN = re.compile(r"^\s*[A-Za-z_]+(?P<yy>\d{2})Q(?P<q>[1-4])\s*$", re.IGNORECASE)
_GDPPLUS_RELEASE_PATTERN = re.compile(r"^\s*GDPPLUS[_-]?(?P<mm>\d{2})(?P<dd>\d{2})(?P<yy>\d{2})\s*$", re.IGNORECASE)


def _coerce_two_digit_year(yy: int) -> int:
    yy = int(yy)
    return 1900 + yy if yy >= 50 else 2000 + yy


def _parse_rtdsm_vintage_period(value: Any) -> Optional[pd.Period]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None

    match = _RTDSM_MONTH_VINTAGE_PATTERN.match(text)
    if match:
        y = _coerce_two_digit_year(match.group('yy'))
        m = int(match.group('m'))
        if 1 <= m <= 12:
            return pd.Period(f"{y:04d}-{m:02d}", freq='M')

    match = _RTDSM_QUARTER_VINTAGE_PATTERN.match(text)
    if match:
        y = _coerce_two_digit_year(match.group('yy'))
        q = int(match.group('q'))
        if 1 <= q <= 4:
            return pd.Period(f"{y}Q{q}", freq='Q')

    return None


def parse_gdpplus_release_timestamp(value: Any) -> Optional[pd.Timestamp]:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, pd.Timestamp):
        return pd.Timestamp(value)

    text = str(value).strip()
    if not text:
        return None

    match = _GDPPLUS_RELEASE_PATTERN.match(text)
    if match:
        y = _coerce_two_digit_year(match.group('yy'))
        m = int(match.group('mm'))
        d = int(match.group('dd'))
        try:
            return pd.Timestamp(year=y, month=m, day=d)
        except Exception:
            return None

    try:
        return pd.Timestamp(pd.to_datetime(text, errors='raise'))
    except Exception:
        return None


def parse_periodish(value: Any, freq_hint: Optional[str] = None) -> Optional[pd.Period]:
    """
    Parse month-like or quarter-like values to a pandas Period without shifting
    to month-end. The parser is conservative about first-of-period timestamps:
    a literal date like ``2026-01-01`` is treated as *January 2026* unless the
    caller explicitly requests quarterly semantics or the label itself encodes a
    quarter (e.g. ``1947:Q1`` or RTDSM-style ``2013:01`` inside GDP/GDPplus
    workbooks).
    """
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, pd.Period):
        return value
    if isinstance(value, pd.Timestamp):
        if freq_hint == 'Q':
            return value.to_period('Q')
        return value.to_period('M')

    text = str(value).strip()
    if text == '':
        return None

    # Explicit quarter encodings used in RTDSM / workbook text.
    for pattern in _QUARTER_PATTERNS:
        match = pattern.match(text)
        if match:
            y = int(match.group('y'))
            q = int(match.group('q'))
            return pd.Period(f"{y}Q{q}", freq='Q')

    match = _RTDSM_NUMERIC_QUARTER_PATTERN.match(text)
    if match:
        y = int(match.group('y'))
        q = int(match.group('q'))
        if 1 <= q <= 4:
            return pd.Period(f"{y}Q{q}", freq='Q')

    # RTDSM vintage headers such as ROUTPUT65M11 / ROUTPUT65Q4.
    vintage_period = _parse_rtdsm_vintage_period(text)
    if vintage_period is not None:
        return vintage_period

    # Handle "YYYYmMM" labels in FRED-QD / FRED-MD filenames and metadata.
    match = re.match(r"^\s*(?P<y>\d{4})[mM](?P<m>\d{2})\s*$", text)
    if match:
        return pd.Period(f"{int(match.group('y')):04d}-{int(match.group('m')):02d}", freq='M')

    # Full timestamps or first-day period markers from the repository. Default to
    # month semantics unless the caller explicitly asks for quarterly semantics.
    try:
        ts = pd.to_datetime(text, errors='raise')
        if freq_hint == 'Q':
            return ts.to_period('Q')
        return ts.to_period('M')
    except Exception:
        return None


def infer_period_frequency_from_values(values: Sequence[Any]) -> Optional[str]:
    """
    Infer whether a sequence is monthly or quarterly while respecting the
    repository's first-day-of-period convention.

    Crucially, the function does **not** infer quarterly frequency merely because
    a date happens to fall on January 1 / April 1 / July 1 / October 1. Quarterly
    status must come from either explicit quarter labels or the sequence pattern
    itself.
    """
    explicit_quarter_hits = 0
    explicit_month_hits = 0
    ordinals: List[int] = []

    for value in values:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            continue
        text = str(value).strip()
        if text == '':
            continue

        if any(pattern.match(text) for pattern in _QUARTER_PATTERNS) or _RTDSM_NUMERIC_QUARTER_PATTERN.match(text):
            explicit_quarter_hits += 1
            p = parse_periodish(text, freq_hint='Q')
            if p is not None:
                ts = p.to_timestamp()
                ordinals.append(ts.year * 12 + ts.month)
            continue

        p_vintage = _parse_rtdsm_vintage_period(text)
        if p_vintage is not None:
            if p_vintage.freqstr.upper().startswith('Q'):
                explicit_quarter_hits += 1
            else:
                explicit_month_hits += 1
            ts = p_vintage.to_timestamp()
            ordinals.append(ts.year * 12 + ts.month)
            continue

        try:
            ts = pd.Timestamp(pd.to_datetime(value, errors='raise'))
            ordinals.append(ts.year * 12 + ts.month)
        except Exception:
            continue

    if explicit_quarter_hits and explicit_quarter_hits >= max(3, explicit_month_hits):
        return 'Q'
    if explicit_month_hits and explicit_month_hits >= max(3, explicit_quarter_hits):
        return 'M'
    if not ordinals:
        return None

    ord_arr = np.unique(np.sort(np.asarray(ordinals, dtype=int)))
    if len(ord_arr) >= 2:
        diffs = np.diff(ord_arr)
        month_set = set((ord_arr - 1) % 12 + 1)
        quarter_like_month_sets = ({1, 4, 7, 10}, {3, 6, 9, 12})
        if len(diffs) > 0:
            if np.mean(diffs == 3) > 0.8 and any(month_set.issubset(s) for s in quarter_like_month_sets):
                return 'Q'
            if np.mean(diffs == 1) > 0.8:
                return 'M'

    months = {(o - 1) % 12 + 1 for o in ord_arr}
    if any(months.issubset(s) for s in ({1, 4, 7, 10}, {3, 6, 9, 12})) and len(ord_arr) >= 4:
        return 'Q'
    return 'M'


def _candidate_excel_tables(path: Path) -> Iterator[Tuple[str, int, pd.DataFrame]]:
    xl = pd.ExcelFile(path)
    for sheet in xl.sheet_names:
        for header in range(0, 12):
            try:
                df = pd.read_excel(path, sheet_name=sheet, header=header)
            except Exception:
                continue
            if df is None or df.empty:
                continue
            df = df.copy()
            df.columns = [normalize_column_name(c).lower() for c in df.columns]
            yield sheet, header, df


def _find_period_column(df: pd.DataFrame) -> Optional[str]:
    best_col = None
    best_score = 0.0
    for col in df.columns:
        score = df[col].map(lambda x: parse_periodish(x, freq_hint='Q') is not None).mean()
        if score > best_score:
            best_score = float(score)
            best_col = col
    return best_col if best_score > 0.4 else None


def _find_year_quarter_columns(df: pd.DataFrame) -> Tuple[Optional[str], Optional[str]]:
    year_col = None
    quarter_col = None
    for col in df.columns:
        low = str(col).lower()
        if year_col is None and low == 'year':
            year_col = col
        if quarter_col is None and low == 'quarter':
            quarter_col = col
    if year_col is None:
        year_matches = [c for c in df.columns if str(c).lower().endswith('year') or str(c).lower() == 'year']
        year_col = year_matches[0] if year_matches else None
    if quarter_col is None:
        quarter_matches = [c for c in df.columns if str(c).lower().endswith('quarter') or str(c).lower() == 'quarter']
        quarter_col = quarter_matches[0] if quarter_matches else None
    return year_col, quarter_col


def _period_from_year_quarter_columns(df: pd.DataFrame) -> Optional[pd.Series]:
    year_col, quarter_col = _find_year_quarter_columns(df)
    if year_col is None or quarter_col is None:
        return None
    year_vals = pd.to_numeric(df[year_col], errors='coerce')
    quarter_vals = pd.to_numeric(df[quarter_col], errors='coerce')
    valid = year_vals.notna() & quarter_vals.isin([1, 2, 3, 4])
    if valid.mean() <= 0.5:
        return None
    out = pd.Series([None] * len(df), index=df.index, dtype='object')
    for idx in df.index[valid]:
        out.loc[idx] = pd.Period(f"{int(year_vals.loc[idx])}Q{int(quarter_vals.loc[idx])}", freq='Q')
    return out


def extract_best_period_matrix(path: Path, required_col_freq: Optional[str] = None) -> Dict[str, Any]:
    """
    Search workbook sheets for a vintage-by-observation matrix where one axis is
    an observation-period axis and the other is a vintage-period axis. This is
    used for RTDSM-style ROUTPUT workbooks.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    candidates: List[Dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        values = _read_sheet_as_values(path, sheet_name)
        values = _trim_2d(values)
        if not values:
            continue

        n_rows = len(values)
        n_cols = max(len(r) for r in values)
        for header_row in range(min(12, n_rows)):
            for index_col in range(min(5, n_cols)):
                cand = _score_matrix_candidate(values, header_row, index_col)
                if cand['score'] < 0:
                    continue
                cand['sheet_name'] = sheet_name
                cand['values'] = values
                if required_col_freq is not None and cand.get('row_freq') != required_col_freq and cand.get('col_freq') != required_col_freq:
                    continue
                candidates.append(cand)

    if not candidates:
        raise ValueError(f"Could not find a date-like matrix in workbook: {path}")

    def ranking_key(c: Dict[str, Any]) -> Tuple[int, int, float]:
        has_required = int(c.get('row_freq') == required_col_freq or c.get('col_freq') == required_col_freq)
        has_monthly_vintage_axis = int(c.get('row_freq') == 'M' or c.get('col_freq') == 'M')
        return (has_required, has_monthly_vintage_axis, float(c['score']))

    best = sorted(candidates, key=ranking_key, reverse=True)[0]
    values = best['values']
    header_row = int(best['header_row'])
    index_col = int(best['index_col'])

    raw_col_labels = values[header_row][index_col + 1:]
    raw_row_labels = [values[i][index_col] if index_col < len(values[i]) else None for i in range(header_row + 1, len(values))]
    row_freq = infer_period_frequency_from_values(raw_row_labels)
    col_freq = infer_period_frequency_from_values(raw_col_labels)
    row_periods = [parse_periodish(v, freq_hint=row_freq) for v in raw_row_labels]
    col_periods = [parse_periodish(v, freq_hint=col_freq) for v in raw_col_labels]

    body = []
    for i in range(header_row + 1, len(values)):
        row = values[i] + [None] * (len(raw_col_labels) + index_col + 1 - len(values[i]))
        body.append([pd.to_numeric(x, errors='coerce') for x in row[index_col + 1:index_col + 1 + len(raw_col_labels)]])
    matrix = pd.DataFrame(body, index=row_periods, columns=col_periods)
    matrix = matrix.dropna(how='all').dropna(axis=1, how='all')

    return {
        'sheet_name': best['sheet_name'],
        'header_row': header_row,
        'index_col': index_col,
        'row_freq': row_freq,
        'col_freq': col_freq,
        'matrix': matrix,
        'raw_row_labels': raw_row_labels,
        'raw_col_labels': raw_col_labels,
    }


def load_routput_vintage_history(path: Path) -> pd.DataFrame:
    matrix_info = extract_best_period_matrix(path, required_col_freq='Q')
    long = melt_vintage_matrix(matrix_info, value_name='level')
    long = long[long['obs_freq'] == 'Q'].copy()
    long = long.dropna(subset=['vintage_period', 'obs_period'])
    long['vintage_timestamp_start'] = long['vintage_period'].map(lambda p: p.to_timestamp() if isinstance(p, pd.Period) else pd.NaT)
    long['quarter_timestamp_start'] = long['obs_period'].map(lambda p: p.to_timestamp() if isinstance(p, pd.Period) else pd.NaT)

    long = long.sort_values(['vintage_period', 'obs_period']).reset_index(drop=True)
    long['gdp_growth_annualized'] = (
        long.groupby('vintage_period')['level']
        .transform(lambda s: 400.0 * np.log(s / s.shift(1)))
    )
    return long


def load_release_truth_table(path: Path) -> pd.DataFrame:
    """
    Load the Philadelphia Fed first/second/third-release workbook into a tidy
    table. The relevant header row contains Date, First, Second, Third,
    Most_Recent.
    """
    best: Optional[pd.DataFrame] = None
    best_score = -np.inf

    for sheet, header, df in _candidate_excel_tables(path):
        period_col = _find_period_column(df)
        if period_col is None:
            continue
        cols_lower = list(df.columns)
        first_col = next((c for c in cols_lower if c == 'first' or c.endswith('_first') or 'first' in c), None)
        second_col = next((c for c in cols_lower if c == 'second' or c.endswith('_second') or 'second' in c), None)
        third_col = next((c for c in cols_lower if c == 'third' or c.endswith('_third') or 'third' in c), None)
        latest_col = next((c for c in cols_lower if 'most_recent' in c or c == 'latest' or 'latest' in c), None)
        required = [first_col, second_col, third_col]
        if any(c is None for c in required):
            continue

        period_share = df[period_col].map(lambda x: parse_periodish(x, freq_hint='Q') is not None).mean()
        numeric_share = pd.concat([
            pd.to_numeric(df[first_col], errors='coerce'),
            pd.to_numeric(df[second_col], errors='coerce'),
            pd.to_numeric(df[third_col], errors='coerce'),
        ], axis=1).notna().mean().mean()
        score = 20.0 * period_share + 20.0 * numeric_share + (5.0 if latest_col is not None else 0.0) - 0.1 * header
        if 'data' in str(sheet).lower():
            score += 2.0
        if score > best_score:
            cols = [period_col, first_col, second_col, third_col] + ([latest_col] if latest_col is not None else [])
            best = df[cols].copy()
            best_score = score

    if best is None:
        raise ValueError(f"Could not detect a release truth table in workbook: {path}")

    period_col = _find_period_column(best)
    assert period_col is not None
    out = pd.DataFrame()
    out['quarter'] = best[period_col].map(lambda x: parse_periodish(x, freq_hint='Q'))
    out['first_release'] = pd.to_numeric(best[[c for c in best.columns if 'first' in str(c).lower()][0]], errors='coerce')
    out['second_release'] = pd.to_numeric(best[[c for c in best.columns if 'second' in str(c).lower()][0]], errors='coerce')
    out['third_release'] = pd.to_numeric(best[[c for c in best.columns if 'third' in str(c).lower()][0]], errors='coerce')
    latest_matches = [c for c in best.columns if 'most_recent' in str(c).lower() or 'latest' in str(c).lower()]
    if latest_matches:
        out['latest'] = pd.to_numeric(best[latest_matches[0]], errors='coerce')

    out = out.dropna(subset=['quarter']).reset_index(drop=True)
    return out


def _extract_wide_quarter_matrix_from_workbook(path: Path, vintage_header_parser) -> Dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    best: Optional[Dict[str, Any]] = None
    best_score = -np.inf

    for sheet_name in wb.sheetnames:
        values = _trim_2d(_read_sheet_as_values(path, sheet_name))
        if not values:
            continue
        n_rows = len(values)
        for header_row in range(min(12, n_rows)):
            row = values[header_row]
            if not row:
                continue
            first_label = normalize_column_name(row[0]).lower() if row[0] is not None else ''
            quarter_labels = [values[i][0] if values[i] else None for i in range(header_row + 1, len(values))]
            parsed_quarters = [parse_periodish(v, freq_hint='Q') for v in quarter_labels]
            release_headers = row[1:]
            parsed_releases = [vintage_header_parser(v) for v in release_headers]
            quarter_count = sum(p is not None for p in parsed_quarters)
            release_count = sum(v is not None for v in parsed_releases)
            if quarter_count < 3 or release_count < 3:
                continue

            body_vals = []
            for i in range(header_row + 1, len(values)):
                current = values[i] + [None] * (len(row) - len(values[i]))
                body_vals.extend([pd.to_numeric(x, errors='coerce') for x in current[1:len(row)]])
            numeric_share = pd.Series(body_vals).notna().mean() if body_vals else 0.0
            score = quarter_count + release_count + 10.0 * numeric_share
            if first_label in {'date', 'quarter'}:
                score += 5.0
            if score > best_score:
                best_score = score
                best = {
                    'sheet_name': sheet_name,
                    'header_row': header_row,
                    'raw_row_labels': quarter_labels,
                    'row_periods': parsed_quarters,
                    'raw_col_labels': release_headers,
                    'col_vintages': parsed_releases,
                    'values': values,
                    'n_cols': len(row),
                }

    if best is None:
        raise ValueError(f"Could not parse wide quarter-vintage workbook: {path}")

    body = []
    for i in range(best['header_row'] + 1, len(best['values'])):
        current = best['values'][i] + [None] * (best['n_cols'] - len(best['values'][i]))
        body.append([pd.to_numeric(x, errors='coerce') for x in current[1:best['n_cols']]])
    matrix = pd.DataFrame(body, index=best['row_periods'], columns=best['col_vintages'])
    matrix = matrix.dropna(how='all').dropna(axis=1, how='all')
    best['matrix'] = matrix
    return best


def load_gdpplus_latest_table(path: Path, value_label: str = 'gdpplus') -> pd.DataFrame:
    matrix_info = _extract_wide_quarter_matrix_from_workbook(path, parse_gdpplus_release_timestamp)
    matrix = matrix_info['matrix'].copy()
    matrix = matrix.loc[~matrix.index.isna(), :]
    if matrix.empty:
        raise ValueError(f"Parsed GDPplus matrix is empty: {path}")

    ordered_cols = sorted(matrix.columns, key=lambda x: pd.Timestamp.min if x is None else x)
    matrix = matrix.reindex(columns=ordered_cols)

    latest_values = []
    latest_release_dates = []
    for _, row in matrix.iterrows():
        notna_cols = [c for c in matrix.columns if pd.notna(row[c])]
        if not notna_cols:
            latest_values.append(np.nan)
            latest_release_dates.append(pd.NaT)
        else:
            last_col = notna_cols[-1]
            latest_values.append(float(row[last_col]))
            latest_release_dates.append(last_col)

    out = pd.DataFrame(
        {
            'quarter': list(matrix.index),
            value_label: latest_values,
            'latest_release_date': latest_release_dates,
        }
    )
    out = out.dropna(subset=['quarter']).reset_index(drop=True)
    return out


def _load_spf_growth_workbook(path: Path, value_label: str) -> pd.DataFrame:
    xl = pd.ExcelFile(path)
    preferred_sheets = [s for s in xl.sheet_names if str(s).strip().upper() == 'RGDP']
    sheet_order = preferred_sheets + [s for s in xl.sheet_names if s not in preferred_sheets]

    best: Optional[pd.DataFrame] = None
    best_score = -np.inf
    chosen_value_col: Optional[str] = None

    for sheet in sheet_order:
        for header in range(0, 8):
            try:
                df = pd.read_excel(path, sheet_name=sheet, header=header)
            except Exception:
                continue
            if df is None or df.empty:
                continue
            df = df.copy()
            df.columns = [normalize_column_name(c).lower() for c in df.columns]
            quarter_series = _period_from_year_quarter_columns(df)
            if quarter_series is None:
                continue

            # Prefer the current-quarter growth forecast in the RGDP sheet.
            preferred_cols = ['drgdp2', 'rgdp2']
            value_col = next((c for c in preferred_cols if c in df.columns), None)
            if value_col is None:
                rgdp_cols = [c for c in df.columns if re.fullmatch(r'd?rgdp[2-6]', str(c).lower())]
                value_col = rgdp_cols[0] if rgdp_cols else None
            if value_col is None:
                continue

            numeric_share = pd.to_numeric(df[value_col], errors='coerce').notna().mean()
            score = 20.0 * quarter_series.notna().mean() + 20.0 * numeric_share - 0.1 * header
            if str(sheet).strip().upper() == 'RGDP':
                score += 5.0
            if score > best_score:
                best_score = score
                best = df.copy()
                best['_quarter_period'] = quarter_series
                chosen_value_col = value_col

    if best is None or chosen_value_col is None:
        raise ValueError(f"Could not parse SPF RGDP growth workbook: {path}")

    out = pd.DataFrame(
        {
            'quarter': best['_quarter_period'],
            value_label: pd.to_numeric(best[chosen_value_col], errors='coerce'),
        }
    )
    out = out.dropna(subset=['quarter']).reset_index(drop=True)
    return out


def load_simple_quarter_value_table(path: Path, value_label: str) -> pd.DataFrame:
    lower_name = path.name.lower()
    if 'gdpplus' in lower_name:
        return load_gdpplus_latest_table(path, value_label=value_label)
    if 'meangrowth' in lower_name or 'mediangrowth' in lower_name or 'meangrowth' in lower_name or 'mediangrowth' in lower_name:
        return _load_spf_growth_workbook(path, value_label=value_label)

    best: Optional[pd.DataFrame] = None
    best_score = -np.inf

    for _, header, df in _candidate_excel_tables(path):
        period_col = _find_period_column(df)
        if period_col is None:
            continue
        num_cols = [c for c in df.columns if c != period_col and pd.to_numeric(df[c], errors='coerce').notna().mean() > 0.5]
        if not num_cols:
            continue
        period_share = df[period_col].map(lambda x: parse_periodish(x, freq_hint='Q') is not None).mean()
        numeric_share = pd.to_numeric(df[num_cols[0]], errors='coerce').notna().mean()
        score = 10.0 * period_share + 10.0 * numeric_share - 0.1 * header
        if score > best_score:
            best = df[[period_col] + num_cols].copy()
            best_score = score

    if best is None:
        raise ValueError(f"Could not parse quarter/value table from workbook: {path}")

    period_col = _find_period_column(best)
    assert period_col is not None
    num_cols = [c for c in best.columns if c != period_col]
    value_col = num_cols[0]

    out = pd.DataFrame(
        {
            'quarter': best[period_col].map(lambda x: parse_periodish(x, freq_hint='Q')),
            value_label: pd.to_numeric(best[value_col], errors='coerce'),
        }
    )
    return out.dropna(subset=['quarter']).reset_index(drop=True)


def build_target_and_truth_objects(repo_root: Path) -> Dict[str, pd.DataFrame]:
    workbooks = select_target_workbooks(repo_root)
    if 'routputMvQd.xlsx' not in workbooks and 'ROUTPUTQvQd.xlsx' not in workbooks:
        raise FileNotFoundError('Could not find ROUTPUT workbook in data/raw.')

    routput_path = workbooks.get('routputMvQd.xlsx', workbooks.get('ROUTPUTQvQd.xlsx'))
    assert routput_path is not None

    target_vintage = load_routput_vintage_history(routput_path)

    truth_objects: Dict[str, pd.DataFrame] = {
        'target_vintage_table': target_vintage
    }

    truth_latest = (
        target_vintage.dropna(subset=['gdp_growth_annualized'])
        .sort_values(['obs_period', 'vintage_period'])
        .groupby('obs_period', as_index=False)
        .tail(1)
        .rename(columns={'obs_period': 'quarter'})
        [['quarter', 'vintage_period', 'gdp_growth_annualized']]
        .rename(columns={'gdp_growth_annualized': 'latest_rtdsm'})
        .reset_index(drop=True)
    )
    truth_objects['truth_latest'] = truth_latest

    release_path = workbooks.get('routput_first_second_third.xlsx')
    if release_path is not None:
        release_truth = load_release_truth_table(release_path)
        truth_objects['truth_release_table'] = release_truth
        if 'third_release' in release_truth.columns:
            truth_objects['truth_third_release'] = (
                release_truth[['quarter', 'third_release']].dropna().reset_index(drop=True)
            )

    gdpplus_path = workbooks.get('GDPplus_Vintages.xlsx')
    if gdpplus_path is not None:
        truth_objects['truth_gdpplus'] = load_simple_quarter_value_table(gdpplus_path, 'gdpplus')

    mean_path = workbooks.get('meanGrowth.xlsx')
    if mean_path is not None:
        truth_objects['spf_mean'] = load_simple_quarter_value_table(mean_path, 'spf_mean')

    median_path = workbooks.get('medianGrowth.xlsx')
    if median_path is not None:
        truth_objects['spf_median'] = load_simple_quarter_value_table(median_path, 'spf_median')

    return truth_objects



def melt_vintage_matrix(matrix_info: Dict[str, Any], value_name: str = 'value') -> pd.DataFrame:
    matrix = matrix_info['matrix'].copy()
    row_freq = matrix_info['row_freq']
    col_freq = matrix_info['col_freq']

    if row_freq not in {'M', 'Q'} and col_freq not in {'M', 'Q'}:
        raise ValueError('Matrix does not expose monthly/quarterly period axes.')

    # Rows are vintage periods, columns are observed periods.
    if row_freq == 'M' and col_freq == 'Q':
        with warnings.catch_warnings():
            warnings.simplefilter('ignore', FutureWarning)
            stacked = matrix.stack(dropna=False)
        long = (
            stacked
            .rename(value_name)
            .rename_axis(index=['vintage_period', 'obs_period'])
            .reset_index()
        )
        long['vintage_freq'] = row_freq
        long['obs_freq'] = col_freq
        return long

    # Rows are observed quarters, columns are vintage months/quarters.
    if row_freq == 'Q' and col_freq in {'M', 'Q'}:
        if col_freq == 'Q':
            row_index = pd.Index([p for p in matrix.index if isinstance(p, pd.Period)])
            col_index = pd.Index([p for p in matrix.columns if isinstance(p, pd.Period)])
            if len(row_index) and len(col_index):
                row_min = row_index.min()
                col_min = col_index.min()
                # In RTDSM quarterly workbooks, observed quarters begin much earlier
                # than vintage quarters. Use that ordering to orient the matrix.
                rows_are_obs = row_min < col_min
            else:
                rows_are_obs = True
        else:
            rows_are_obs = True

        if rows_are_obs:
            with warnings.catch_warnings():
                warnings.simplefilter('ignore', FutureWarning)
                stacked = matrix.T.stack(dropna=False)
            long = (
                stacked
                .rename(value_name)
                .rename_axis(index=['vintage_period', 'obs_period'])
                .reset_index()
            )
            long['vintage_freq'] = col_freq
            long['obs_freq'] = row_freq
            return long

    raise ValueError('Could not orient matrix as vintage x observed-period.')
